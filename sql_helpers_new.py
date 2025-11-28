# -*- coding: utf-8 -*-

import math
import pyodbc
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.pagebreak import PageBreak
from tkinter import messagebox
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles import Font
from datetime import datetime
from tkinter import filedialog
import os
import platform
import subprocess
from copy import copy
from io import BytesIO
import shutil
from decimal import Decimal, InvalidOperation
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='pandas')
warnings.filterwarnings(
    'ignore',
    message='pandas only supports SQLAlchemy connectable',
    category=UserWarning,
)
import pandas as pd

_template_cache = {}


def _get_template_stream(template_path):
    """
    Return a fresh in-memory stream for the given Excel template.
    The binary contents are cached after the first load to avoid extra disk I/O.
    """
    cache_key = os.path.abspath(template_path)
    data = _template_cache.get(cache_key)
    if data is None:
        with open(template_path, 'rb') as f:
            data = f.read()
        _template_cache[cache_key] = data
    return BytesIO(data)


def _format_sotk_value(value):
    """Chuẩn hóa số tờ khai để dùng làm tên file (loại bỏ .0, khoảng trắng)."""
    def _clean(text):
        if text is None:
            return ''
        txt = str(text).strip()
        if not txt:
            return ''
        if txt.endswith('.0'):
            try:
                return str(int(float(txt)))
            except Exception:
                stripped = txt.rstrip('0').rstrip('.')
                return stripped or txt
        return txt

    normalized = None
    try:
        normalized = normalize_dtokhaimdid(value)
    except Exception:
        normalized = None
    if normalized:
        return _clean(normalized)
    try:
        return _clean(str(int(float(value))))
    except Exception:
        return _clean(value)


def normalize_dtokhaimdid(value):
    """
    Ensure declaration IDs keep a stable string form so lookups work even if the
    database driver returns them as floats like 12345.0.
    """
    if value is None:
        return ''
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()

    text = str(value).strip()
    if not text:
        return ''
    if text.endswith('.0') and text[:-2].isdigit():
        return text[:-2]
    try:
        dec = Decimal(text.replace(',', '.'))
        if dec == dec.to_integral_value():
            return str(dec.quantize(Decimal('1')))
    except InvalidOperation:
        pass
    return text
def get_connection(Sqlhost):
    (host, database, user, password) = Sqlhost
    conn_str = f'''DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={host};DATABASE={database};UID={user};PWD={password};Encrypt=yes;TrustServerCertificate=yes;'''
    return pyodbc.connect(conn_str)

def execute_query(Sqlhost, query):
    conn = get_connection(Sqlhost)
    try:
        df = pd.read_sql_query(query, conn)
    finally:
        conn.close()
    return df

def select_query_df_pyodbc(Sqlhost, query, params=None):
    conn = get_connection(Sqlhost)
    try:
        cur = conn.cursor()
        if params is None:
            cur.execute(query)
        else:
            cur.execute(query, params)
        rows = cur.fetchall()
        columns = [col[0] for col in cur.description] if cur.description else []
        df = pd.DataFrame.from_records(rows, columns=columns)
        return df
    finally:
        conn.close()

def SQLrun_query(Sqlhost, query):
    conn = get_connection(Sqlhost)
    cursor = conn.cursor()
    cursor.execute(query)
    conn.commit()
    cursor.close()
    conn.close()

def SQLrun_query_params(Sqlhost, query, params):
    conn = get_connection(Sqlhost)
    cursor = conn.cursor()
    cursor.execute(query, params)
    conn.commit()
    cursor.close()
    conn.close()

def SQL_Select_query_params(Sqlhost, query, params=None):
    conn = get_connection(Sqlhost)
    try:
        if params is None:
            df = pd.read_sql_query(query, conn)
        else:
            df = pd.read_sql_query(query, conn, params=params)
    finally:
        conn.close()
    return df
def clear_all_page_breaks(worksheet):
    try:
        from openpyxl.worksheet.pagebreak import PageBreak
        worksheet.row_breaks = PageBreak()
        worksheet.col_breaks = PageBreak()
    except:
        pass
    
    for row_idx, row_dimension in worksheet.row_dimensions.items():
        if hasattr(row_dimension, 'page_break'):
            row_dimension.page_break = False
    
    for col_idx, col_dimension in worksheet.column_dimensions.items():
        if hasattr(col_dimension, 'page_break'):
            col_dimension.page_break = False
    
    # Phương pháp 3: Thử tạo mới đối tượng PageBreaks
    try:
        from openpyxl.worksheet.page import PageBreak
        worksheet.row_breaks = PageBreak()
        worksheet.col_breaks = PageBreak()
    except:
        pass
def get_cd_details_df(Sqlhost, dtokhaimdid):
    queries = {
        'dtokhaimd': (
            """
            SELECT *
            FROM DTOKHAIMD
            WHERE _DTOKHAIMDID = ?
            """,
            (dtokhaimdid,)
        ),
        'vnaccs': (
            """
            SELECT *
            FROM DTOKHAIMD_VNACCS
            WHERE _DTOKHAIMDID = ?
            """,
            (dtokhaimdid,)
        ),
        'vnaccs2': (
            """
            SELECT *
            FROM DTOKHAIMD_VNACCS2
            WHERE _DTOKHAIMDID = ?
            """,
            (dtokhaimdid,)
        )
    }
    
    result = {}
    conn = None
    
    try:
        conn = get_connection(Sqlhost)
        
        for key, (query, params) in queries.items():
            try:
                df = pd.read_sql_query(query, conn, params=params)
                result[key] = df
                
                if df.empty:
                    print(f"Đã lấy thành công {len(df)} bản ghi từ bảng {key}")
                #else:
                    print(f"Không tìm thấy dữ liệu trong bảng {key} với ID: {dtokhaimdid}")
                    
            except Exception as e:
                error_msg = f'Lỗi khi lấy dữ liệu từ bảng {key}: {str(e)}'
                print(f"Warning: {error_msg}")
                messagebox.showwarning('DB Warning', error_msg)
                result[key] = pd.DataFrame()
                
    except Exception as e:
        error_msg = f'Lỗi kết nối database: {str(e)}'
        print(f"Error: {error_msg}")
        messagebox.showerror('DB Error', error_msg)
        result = {key: pd.DataFrame() for key in queries.keys()}
        
    finally:
        if conn:
            conn.close()
    
    return result


def get_cd_details_bulk(Sqlhost, dtokhaimdid_list):
    """
    Batch version of get_cd_details_df that fetches data for multiple declarations
    in a single connection round-trip per table.
    """
    if not dtokhaimdid_list:
        return {}

    seen = set()
    lookup_keys = []
    sql_params = []

    for raw_id in dtokhaimdid_list:
        text_id = normalize_dtokhaimdid(raw_id)
        if not text_id or text_id in seen:
            continue
        seen.add(text_id)
        lookup_keys.append(text_id)
        try:
            sql_params.append(int(text_id))
        except (ValueError, TypeError):
            sql_params.append(text_id)

    if not lookup_keys:
        return {}

    placeholders = ','.join(['?'] * len(sql_params))
    result = {
        key: {
            'dtokhaimd': pd.DataFrame(),
            'vnaccs': pd.DataFrame(),
            'vnaccs2': pd.DataFrame(),
            'dhangmddk': pd.DataFrame(),
        } for key in lookup_keys
    }

    queries = {
        'dtokhaimd': f"""
            SELECT *
            FROM DTOKHAIMD
            WHERE _DTOKHAIMDID IN ({placeholders})
        """,
        'vnaccs': f"""
            SELECT *
            FROM DTOKHAIMD_VNACCS
            WHERE _DTOKHAIMDID IN ({placeholders})
        """,
        'vnaccs2': f"""
            SELECT *
            FROM DTOKHAIMD_VNACCS2
            WHERE _DTOKHAIMDID IN ({placeholders})
        """,
        'dhangmddk': f"""
            SELECT *
            FROM DHANGMDDK
            WHERE _DTOKHAIMDID IN ({placeholders})
        """,
    }

    conn = None
    try:
        conn = get_connection(Sqlhost)
        params_tuple = tuple(sql_params)
        for key, query in queries.items():
            try:
                df = pd.read_sql_query(query, conn, params=params_tuple)
            except Exception as exc:
                error_msg = f'L��-i khi l���y d��_ li���u t��� {key}: {exc}'
                print(f"Warning: {error_msg}")
                messagebox.showwarning('DB Warning', error_msg)
                continue

            if df.empty:
                continue
            id_column = None
            for col in df.columns:
                if isinstance(col, str) and col.strip().lower() == '_dtokhaimdid':
                    id_column = col
                    break

            if id_column is None:
                error_msg = (
                    f"L��-i: b���ng {key} khA'ng cA3 cA?t '_DTOKHAIMDID'. "
                    f"Cac cot thu duoc: {list(df.columns)}"
                )
                print(f"Warning: {error_msg}")
                messagebox.showwarning('DB Warning', error_msg)
                continue

            for _id, group in df.groupby(id_column):
                str_id = normalize_dtokhaimdid(_id)
                if str_id not in result:
                    result[str_id] = {
                        'dtokhaimd': pd.DataFrame(),
                        'vnaccs': pd.DataFrame(),
                        'vnaccs2': pd.DataFrame(),
                        'dhangmddk': pd.DataFrame(),
                    }
                result[str_id][key] = group.reset_index(drop=True)
    except Exception as exc:
        error_msg = f'L��-i k���t n��`i database (batch): {exc}'
        print(f"Error: {error_msg}")
        messagebox.showerror('DB Error', error_msg)
        return {}
    finally:
        if conn:
            conn.close()

    return result

# ========== EXCEL HELPERS FUNCTIONS ==========

def format_date(val):
    """Định dạng ngày về dd/mm/YYYY"""
    if isinstance(val, datetime):
        return val.strftime('%d/%m/%Y')
    if isinstance(val, str):
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%m-%d-%Y'):
            try:
                return datetime.strptime(val.strip(), fmt).strftime('%d/%m/%Y')
            except Exception:
                pass
    return val

def to_int(val):
    """Chuyển sang số nguyên"""
    return int(val)

def copy_cell_format(src_cell, dest_cell):
    """Sao chép giá trị và style của 1 ô sang ô khác"""
    dest_cell.value = src_cell.value
    if src_cell.has_style:
        dest_cell.font = copy(src_cell.font)
        dest_cell.fill = copy(src_cell.fill)
        dest_cell.border = copy(src_cell.border)
        dest_cell.alignment = copy(src_cell.alignment)
        dest_cell.number_format = copy(src_cell.number_format)
        dest_cell.protection = copy(src_cell.protection)

def copy_row_range_with_style(src_ws, dest_ws, cell_range, start_row_dest):
    """Sao chép cả một vùng hàng kèm style từ sheet nguồn sang sheet đích"""
    (min_col, min_row, max_col, max_row) = range_boundaries(cell_range)
    row_offset = start_row_dest - min_row
    
    for row in src_ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            new_row = cell.row + row_offset
            new_col = cell.column
            dest_cell = dest_ws.cell(row=new_row, column=new_col)
            copy_cell_format(cell, dest_cell)
    
    for merged_range in src_ws.merged_cells.ranges:
        if (merged_range.min_row >= min_row and merged_range.max_row <= max_row and 
            merged_range.min_col >= min_col and merged_range.max_col <= max_col):
            dest_ws.merge_cells(start_row=merged_range.min_row + row_offset, 
                               start_column=merged_range.min_col, 
                               end_row=merged_range.max_row + row_offset, 
                               end_column=merged_range.max_col)

TKN_DETAIL_FIELD_MAP = {
    'STTHANG': ('C10',),
    'MA_HANGKB': ('G11',),
    'TEN_HANG': ('G12',),
    'MA_HANG_QL_RIENG': ('Q11',),
    'LUONG': ('V15',),
    'LUONG2': ('V16',),
    'THUE_XNK': ('I22',),
    'TRIGIA_HDTM': ('I17',),
    'DGIA_HDTM': ('V17',),
    'MA_NT_DGIA_HDTM': ('AC17',),
    'DVT_DGIA_HDTM': ('AE15', 'AE17', 'AE20'),
    'TRIGIA_TT_S': ('I19',),
    'DGIA_TT': ('V20',),
    'MA_PL_TS_TNK': ('H21',),
    'TEN_TS_TNK': ('I21',),
    'NUOC_XX': ('X22',),
    'TEN_NUOC_XX': ('Z22',),
    'TS_XNK_MA_BT': ('AC22',),
    'MIENTHUE_THUE_GIAM': ('I23',),
    'MIENTHUE_MA': ('E27',),
    'DK_MIEN_GIAM_THUE': ('I27',),
    'THUE_VAT': ('I32',),
    'THUEKHAC_TEN_KHOAN_MUC': ('H29',),
    'THUEKHAC_MA_AP_DUNG': ('W29',),
    'THUEKHAC_TRGIA_TT': ('I30',),
    'THUEKHAC_LUONG_TT': ('W30',),
    'THUEKHAC_TEN_TS': ('I31',),
    'THUEKHAC_SO_TIEN': ('I32',),
    'THUEKHAC_DK_MIEN_GIAM_THUE': ('U32',),
    'THUEKHAC_MA_MIEN_GIAM': ('S32',),
    'THUEKHAC_THUE_GIAM': ('I33',),
    'THUEKHAC_TEN_KHOAN_MUC2': ('H34',),
    'THUEKHAC_MA_AP_DUNG2': ('W34',),
    'THUEKHAC_TRGIA_TT2': ('I35',),
    'THUEKHAC_LUONG_TT2': ('W35',),
    'THUEKHAC_TEN_TS2': ('I36',),
    'THUEKHAC_SO_TIEN2': ('I37',),
    'THUEKHAC_DK_MIEN_GIAM_THUE2': ('U37',),
    'THUEKHAC_MA_MIEN_GIAM2': ('S37',),
    'THUEKHAC_THUE_GIAM2': ('I38',),
    'THUEKHAC_TEN_KHOAN_MUC3': ('H39',),
    'THUEKHAC_MA_AP_DUNG3': ('W39',),
    'THUEKHAC_TRGIA_TT3': ('I40',),
    'THUEKHAC_LUONG_TT3': ('W40',),
    'THUEKHAC_TEN_TS3': ('I41',),
    'THUEKHAC_SO_TIEN3': ('I42',),
    'THUEKHAC_DK_MIEN_GIAM_THUE3': ('U42',),
    'THUEKHAC_MA_MIEN_GIAM3': ('S42',),
    'THUEKHAC_THUE_GIAM3': ('I43',),
    'THUEKHAC_TEN_KHOAN_MUC4': ('H44',),
    'THUEKHAC_MA_AP_DUNG4': ('W44',),
    'THUEKHAC_TRGIA_TT4': ('I45',),
    'THUEKHAC_LUONG_TT4': ('W45',),
    'THUEKHAC_TEN_TS4': ('I46',),
    'THUEKHAC_SO_TIEN4': ('I47',),
    'THUEKHAC_DK_MIEN_GIAM_THUE4': ('U47',),
    'THUEKHAC_MA_MIEN_GIAM4': ('S47',),
    'THUEKHAC_THUE_GIAM4': ('I48',),
    'THUEKHAC_TEN_KHOAN_MUC5': ('H49',),
    'THUEKHAC_MA_AP_DUNG5': ('W49',),
    'THUEKHAC_TRGIA_TT5': ('I50',),
    'THUEKHAC_LUONG_TT5': ('W50',),
    'THUEKHAC_TEN_TS5': ('I51',),
    'THUEKHAC_SO_TIEN5': ('I52',),
    'THUEKHAC_DK_MIEN_GIAM_THUE5': ('U52',),
    'THUEKHAC_MA_MIEN_GIAM5': ('S52',),
    'MA_DVT2': ('AE16',),
    
}

TKN_DETAIL_TEMPLATE_RANGE = "A1:AH53"
TKN_DETAIL_PAGE_HEIGHT = 53
TKN_DETAIL_FIRST_DETAIL_ROW = 139
TKN_DETAIL_CLEAR_CELLS = tuple(
    cell
    for cells in TKN_DETAIL_FIELD_MAP.values()
    for cell in cells
)

TKN_DETAIL_NUMERIC_FIELDS = {
    'LUONG',
    'LUONG2',
    'TRIGIA_HDTM',
    'DGIA_HDTM',
    'DGIA_TT',
    'TRIGIA_TT_S',
    'MIENTHUE_THUE_GIAM',
    'THUEKHAC_TRGIA_TT',
    'THUEKHAC_LUONG_TT',
    'THUEKHAC_SO_TIEN',
    'THUEKHAC_TRGIA_TT2',
    'THUEKHAC_LUONG_TT2',
    'THUEKHAC_SO_TIEN2',
    'THUEKHAC_TRGIA_TT3',
    'THUEKHAC_LUONG_TT3',
    'THUEKHAC_SO_TIEN3',
    'THUEKHAC_TRGIA_TT4',
    'THUEKHAC_LUONG_TT4',
    'THUEKHAC_SO_TIEN4',
    'THUEKHAC_TRGIA_TT5',
    'THUEKHAC_LUONG_TT5',
    'THUEKHAC_SO_TIEN5',
}

TKN_DETAIL_NUMERIC_CELLS = {
    'V15',
    'V16',
    'V17',
    'V20',
    'I17',
    'I22',
    'I23',
    'I30',
    'I32',
    'I35',
    'I37',
    'I40',
    'I42',
    'I45',
    'I47',
    'I50',
    'I52',
    'W30',
    'W35',
    'W40',
    'W45',
    'W50',
}

TKX_DETAIL_ODD_FIELD_MAP = {
    'STTHANG': ('C11',),
    'MA_HANGKB': ('F13',),
    'TEN_HANG': ('F14',),
    'MA_DVT': ('Y17',),
    'LUONG': ('Q17',),
    'DGIA_TT': ('R22',),
    'TRIGIA_TT': ('G21',),
    'MA_NT_THUE_XNK': ('L24',),
    'LUONG2': ('Q18',),
    'MA_DVT2': ('Y18',),
    'TRIGIA_HDTM': ('F19',),
    'DGIA_HDTM': ('R19',),
    'MA_NT_DGIA_HDTM': ('W19',),
    'DVT_DGIA_HDTM': ('Y19',),
    'MA_NT_TRIGIA_TT_S': ('M21',),
    'MA_NT_DGIA_TT': ('Y22',),
    'DVT_DGIA_TT': ('AA22',),
}

TKX_DETAIL_EVEN_FIELD_MAP = {
    'STTHANG': ('C35',),
    'MA_HANGKB': ('F37',),
    'TEN_HANG': ('F38',),
    'MA_DVT': ('Y41',),
    'LUONG': ('Q41',),
    'DGIA_TT': ('R46',),
    'TRIGIA_TT': ('G45',),
    'MA_NT_THUE_XNK': ('L48',),
    'LUONG2': ('Q42',),
    'MA_DVT2': ('Y42',),
    'TRIGIA_HDTM': ('F43',),
    'DGIA_HDTM': ('R43',),
    'MA_NT_DGIA_HDTM': ('W43',),
    'DVT_DGIA_HDTM': ('Y43',),
    'MA_NT_TRIGIA_TT_S': ('M45',),
    'MA_NT_DGIA_TT': ('Y46',),
    'DVT_DGIA_TT': ('AA46',),
}

TKX_DETAIL_TEMPLATE_FULL_RANGE = "A1:AC57"
TKX_DETAIL_TEMPLATE_HALF_RANGE = "A1:AC34"
TKX_DETAIL_PAGE_HEIGHT = 57
TKX_DETAIL_PARTIAL_HEIGHT = 34
TKX_DETAIL_FIRST_DETAIL_ROW = 145
TKX_DETAIL_HEADER_ROWS = 144

TKX_DETAIL_ODD_CLEAR_CELLS = tuple(
    cell
    for cells in TKX_DETAIL_ODD_FIELD_MAP.values()
    for cell in cells
)
TKX_DETAIL_EVEN_CLEAR_CELLS = tuple(
    cell
    for cells in TKX_DETAIL_EVEN_FIELD_MAP.values()
    for cell in cells
)

TKX_DETAIL_FORCE_FONT_CELLS = {
    'Y17',
    'Y18',
    'Y19',
    'W19',
    'AA22',
    'Y41',
    'Y42',
    'Y43',
    'W43',
    'AA46',
}


def clear_cells_by_mapping(worksheet, cells, safe_writer):
    """Dat gia tri rong cho danh sach o dua tren mapping."""
    for cell_address in cells:
        safe_writer(worksheet, cell_address, '')


def format_tkn_detail_value(field, value, cell_address):
    """�`A�nh dA�ng giA� tr��< tr��c khi ghi vA�o sheet chi tiA�t TKN."""
    if value is None:
        return ''
    if isinstance(value, str):
        cleaned_value = value.strip()
    else:
        cleaned_value = value

    if cleaned_value == '':
        return ''

    if field == 'STTHANG':
        text = str(cleaned_value).strip()
        if not text:
            return ''
        try:
            normalized = int(float(text.replace(',', '.')))
            return f"<{normalized:02d}>"
        except (ValueError, TypeError):
            if text.isdigit():
                try:
                    normalized = int(text)
                    return f"<{normalized:02d}>"
                except Exception:
                    pass
            return f"<{text}>"

    if field in TKN_DETAIL_NUMERIC_FIELDS or cell_address in TKN_DETAIL_NUMERIC_CELLS:
        formatted = format_number_vn(cleaned_value)
        if formatted in (None, ''):
            return ''
        return str(formatted)

    if isinstance(cleaned_value, str):
        return cleaned_value
    return str(cleaned_value)


def populate_worksheet_from_mapping(worksheet, row, mapping, safe_writer):
    """Ghi du lieu tu Series dua tren mapping o -> cot."""
    if row is None:
        return
    for field, cells in mapping.items():
        if isinstance(row, dict):
            value = row.get(field, None)
        elif hasattr(row, 'index') and field in row.index:
            value = row[field]
        elif hasattr(row, field):
            value = getattr(row, field)
        else:
            value = None

        if value is None:
            continue

        if isinstance(value, str):
            value = value.strip()
            if value == '':
                continue
        else:
            try:
                if pd.isna(value):
                    continue
            except Exception:
                pass

        for cell_address in cells:
            formatted = format_tkn_detail_value(field, value, cell_address)
            if formatted == '':
                continue
            safe_writer(worksheet, cell_address, formatted)

def format_number(value, decimal_places=None):
    """Định dạng số với số chữ số thập phân"""
    try:
        number = float(value)
    except Exception:
        return value
    if isinstance(decimal_places, int) and decimal_places >= 0:
        return f"{number:.{decimal_places}f}"
    return number

def format_number_vn(value):
    """
    Định dạng số theo kiểu Việt Nam
    Ví dụ: 27982 -> 27.982
           62073.1 -> 62.073,1
           1624515100.1 -> 1.624.515.100,1
           88749891 -> 88.749.891 (không có ,0)
    """
    if value is None or value == '':
        return value
    
    try:
        # Chuyển sang float để xử lý
        num = float(value)
        
        # Kiểm tra nếu là số nguyên
        if num == int(num):
            # Nếu là số nguyên, chỉ định dạng phần nguyên
            integer_part = str(int(num))
            formatted_integer = f"{int(integer_part):,}".replace(',', '.')
            return formatted_integer
        else:
            # Nếu có phần thập phân, xử lý như cũ
            if '.' in str(num):
                integer_part, decimal_part = str(num).split('.')
            else:
                integer_part = str(int(num))
                decimal_part = None
            
            # Định dạng phần nguyên với dấu chấm làm phân cách hàng nghìn
            formatted_integer = f"{int(integer_part):,}".replace(',', '.')
            
            # Nếu có phần thập phân, thêm vào với dấu phẩy
            if decimal_part:
                return f"{formatted_integer},{decimal_part}"
            else:
                return formatted_integer
                
    except (ValueError, TypeError):
        return value

def get_first_sheet_name(file_path):
    """Lấy tên sheet đầu tiên trong file Excel"""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.xlsx':
        wb = openpyxl.load_workbook(file_path, read_only=True)
        if wb.sheetnames:
            return wb.sheetnames[0]
        return None
    if ext == '.xls':
        try:
            import xlrd
        except Exception:
            return None
        wb = xlrd.open_workbook(file_path)
        names = wb.sheet_names()
        if names:
            return names[0]
        return None
    return ''

def Excel_data(filepath, sheetname, col_start, col_end, date_cols):
    """Đọc dữ liệu từ Excel theo cột và hàng quy ước"""
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheetname]
    labels = {}
    
    for col in range(col_start, col_end + 1):
        col_letter = get_column_letter(col).lower()
        labels[col_letter] = ws.cell(5, col).value
    
    data = []
    max_row = ws.max_row
    
    for row_idx in range(6, max_row + 1):
        row_values = []
        empty_row = True
        
        for col in range(col_start, col_end + 1):
            val = ws.cell(row_idx, col).value
            
            if date_cols:
                is_date_col = False
                if isinstance(date_cols, (list, tuple, set)):
                    is_date_col = (col in date_cols) or (get_column_letter(col).lower() in {str(c).lower() for c in date_cols})
                elif isinstance(date_cols, int):
                    is_date_col = (col == date_cols)
                elif isinstance(date_cols, str):
                    is_date_col = (get_column_letter(col).lower() == date_cols.lower())
                
                if is_date_col and val is not None and val != '':
                    val = format_date(val)
            
            if val not in (None, ''):
                empty_row = False
            row_values.append(val)
        
        if empty_row:
            continue
        data.append(row_values)
    
    return labels, data

def write_to_excel(data, filepath, sheetname):
    """Ghi dữ liệu dạng mảng 2 chiều vào sheet"""
    wb = load_workbook(filepath)
    
    if sheetname in wb.sheetnames:
        ws = wb[sheetname]
        for row in ws.iter_rows(min_row=8, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet(sheetname)
    
    start_row = 8
    start_col = 3
    
    for i, row in enumerate(data):
        for j, val in enumerate(row):
            ws.cell(row=start_row + i, column=start_col + j, value=val)
    
    wb.save(filepath)

def export_treeview_to_excel(treeview, default_filename='Export.xlsx'):
    """Xuất dữ liệu từ Treeview ra file Excel mới và mở sau khi lưu"""
    file_path = filedialog.asksaveasfilename(
        defaultextension='.xlsx', 
        initialfile=default_filename, 
        filetypes=[('Excel files', '*.xlsx')]
    )
    
    if not file_path:
        return False
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Export'
    columns = treeview['columns']
    
    # Ghi header
    for col_idx, col in enumerate(columns, start=1):
        header_text = treeview.heading(col).get('text', col)
        ws.cell(row=1, column=col_idx, value=header_text)
    
    # Ghi từng dòng dữ liệu
    for row_idx, item_id in enumerate(treeview.get_children(''), start=2):
        values = treeview.item(item_id, 'values')
        for col_idx, cell_val in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell_val)
    
    wb.save(file_path)
    
    try:
        if platform.system() == 'Windows':
            os.startfile(file_path)
        elif platform.system() == 'Darwin':
            subprocess.call(['open', file_path])
        else:
            subprocess.call(['xdg-open', file_path])
    except Exception:
        pass
    
    return True

def read_excel_range(file_path, sheet_name, start_col, end_col='M'):
    """Đọc một khoảng dữ liệu từ sheet theo cột bắt đầu-kết thúc"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    
    last_row = ws.max_row
    for r in range(ws.max_row, 1, -1):
        val = ws[f'A{r}'].value
        if val not in (None, ''):
            last_row = r
            break
    
    def col_to_idx(c):
        if isinstance(c, int):
            return c
        return openpyxl.utils.column_index_from_string(c)
    
    start_idx = col_to_idx(start_col)
    end_idx = col_to_idx(end_col)
    
    if end_idx < start_idx:
        start_idx, end_idx = end_idx, start_idx
    
    data = []
    for r in range(1, last_row + 1):
        row_vals = []
        empty_row = True
        
        for c in range(start_idx, end_idx + 1):
            v = ws.cell(r, c).value
            if v not in (None, ''):
                empty_row = False
            row_vals.append(v)
        
        if empty_row and r > 1:
            continue
        data.append(row_vals)
    
    return data

def load_excel_to_dataframe(file_path, sheet_name, start_row, start_col=0):
    """Đọc Excel thành DataFrame, cắt theo vị trí bắt đầu hàng/cột"""
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    df = df.iloc[start_row:, start_col:]
    df = df.reset_index(drop=True)
    
    if not df.empty:
        df.columns = df.iloc[0]
        df = df[1:].reset_index(drop=True)
    
    df = df.fillna('')
    return df

def PrintCD_TKX(dtokhaimdid, index, data, output_path):
    """In/ghi Tờ khai nhập (TKN) dựa vào template và dữ liệu"""
    os.makedirs(output_path, exist_ok=True)
    template_path = 'Excel_Tempt/TKX_TEMPT.xlsx'
    wb = load_workbook(_get_template_stream(template_path))
    ws1 = wb['TKX_Tempt1']
    ws2 = wb['TKX_Tempt2']
    tkx_font_templates = {}
    for addr in TKX_DETAIL_FORCE_FONT_CELLS:
        try:
            cell = ws2[addr]
            if cell is not None and cell.has_style and cell.font is not None:
                tkx_font_templates[addr] = copy(cell.font)
            else:
                tkx_font_templates[addr] = Font(name='Courier New', size=8)
        except Exception:
            tkx_font_templates[addr] = Font(name='Courier New', size=8)
    
    # Lấy dữ liệu từ các bảng
    dtokhaimd_row = None
    vnaccs_row = None
    vnaccs2_row = None
    
    if not data['dtokhaimd'].empty:
        dtokhaimd_row = data['dtokhaimd'].iloc[0]
    if not data['vnaccs'].empty:
        vnaccs_row = data['vnaccs'].iloc[0]
    if not data['vnaccs2'].empty:
        vnaccs2_row = data['vnaccs2'].iloc[0]

    sotk_value = _format_sotk_value(dtokhaimdid)
    if dtokhaimd_row is not None and 'SOTK' in dtokhaimd_row:
        candidate = _format_sotk_value(dtokhaimd_row['SOTK'])
        if candidate:
            sotk_value = candidate
    output_file = os.path.join(output_path, f'ToKhaiHQ7N_QDTQ_{sotk_value}.xlsx')
    
    total_pages_declared = None
    if dtokhaimd_row is not None and 'SO_TRANG_TK' in dtokhaimd_row:
        try:
            total_pages_declared = int(dtokhaimd_row['SO_TRANG_TK'])
        except (ValueError, TypeError):
            total_pages_declared = None
    
    total_pages_declared = None
    if dtokhaimd_row is not None and 'SO_TRANG_TK' in dtokhaimd_row:
        try:
            total_pages_declared = int(dtokhaimd_row['SO_TRANG_TK'])
        except (ValueError, TypeError):
            total_pages_declared = None
    
    def restore_tkx_font_if_needed(worksheet, cell_address, cell_obj):
        if cell_obj is None:
            return
        if worksheet is ws2 and cell_address in tkx_font_templates:
            tmpl_font = tkx_font_templates.get(cell_address)
            if tmpl_font is not None:
                try:
                    cell_obj.font = copy(tmpl_font)
                except Exception:
                    cell_obj.font = Font(name='Courier New', size=8)

    def safe_set_cell(worksheet, cell_address, value):
        """Ghi giá trị vào ô một cách an toàn, xử lý merged cells"""
        try:
            cell = worksheet[cell_address]
            target_cell = cell
            if isinstance(cell, MergedCell):
                target = None
                for merged_range in worksheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        target = worksheet.cell(merged_range.min_row, merged_range.min_col)
                        break
                if target is None:
                    row = getattr(cell, 'row', None)
                    col = getattr(cell, 'col_idx', None) or getattr(cell, 'column', None)
                    if row is not None and col is not None:
                        target = worksheet.cell(row=row, column=col)
                target_cell = target or cell
                target_cell.value = value
            else:
                cell.value = value
            restore_tkx_font_if_needed(worksheet, cell_address, target_cell if isinstance(cell, MergedCell) else cell)
        except Exception as e:
            print(f"Warning: Không thể ghi vào ô {cell_address}: {e}")
    
    # Điền dữ liệu vào ws1 theo mapping
    try:
        # Từ bảng DTOKHAIMD
        if dtokhaimd_row is not None:
            # Thông tin cơ bản
            if 'SOTK' in dtokhaimd_row:
                safe_set_cell(ws1, 'E4', dtokhaimd_row['SOTK'])
                safe_set_cell(ws2, 'E4', dtokhaimd_row['SOTK'])
            trangthaitokhai = 'Tờ khai hàng hóa xuất khẩu (thông quan)'
            if 'SOTK' in dtokhaimd_row:
                safe_set_cell(ws1, 'E4', dtokhaimd_row['SOTK'])
                safe_set_cell(ws2, 'E4', dtokhaimd_row['SOTK'])
            #if 'MA_LH' in dtokhaimd_row:
            if dtokhaimd_row['TTTK'] == 'P':
                trangthaitokhai = 'Tờ khai hàng hóa xuất khẩu (thông báo kết quả phân luồng)'
                safe_set_cell(ws1, 'F2', trangthaitokhai)
                safe_set_cell(ws2, 'D2', trangthaitokhai)    #safe_set_cell(ws1, 'P6', dtokhaimd_row['MA_LH'])
            elif dtokhaimd_row['TTTK'] == 'T':
                trangthaitokhai = 'Tờ khai hàng hóa xuất khẩu (thông quan)'
                safe_set_cell(ws1, 'F2', trangthaitokhai)
                safe_set_cell(ws2, 'D2', trangthaitokhai)
            elif dtokhaimd_row['TTTK'] == 'E':
                trangthaitokhai = 'Bản xác nhận nội dung tờ khai hàng hóa xuất khẩu<In thử>'
                safe_set_cell(ws1, 'F2', trangthaitokhai)
                safe_set_cell(ws2, 'D2', trangthaitokhai)
            if 'MA_LH' in dtokhaimd_row:
                safe_set_cell(ws1, 'L6', dtokhaimd_row['MA_LH'])
                
            if 'SOTK_DAU_TIEN' in dtokhaimd_row:
                safe_set_cell(ws1, 'P4', dtokhaimd_row['SOTK_DAU_TIEN'])
                safe_set_cell(ws2, 'P4', dtokhaimd_row['SOTK_DAU_TIEN'])
            
            if 'SOTK_NHANH' in dtokhaimd_row and dtokhaimd_row['SOTK_NHANH'] is not None and str(dtokhaimd_row['SOTK_NHANH']).strip() != '':
                tknhanh = f"{dtokhaimd_row['SOTK_NHANH']}/{dtokhaimd_row['SOTK_TONG']}"
                safe_set_cell(ws1, 'U4', tknhanh)
                safe_set_cell(ws2, 'U4', tknhanh)
            Ma_nt_value = dtokhaimd_row['MA_NT_TGTT']
            if 'MA_NT_TGTT' in dtokhaimd_row:
                safe_set_cell(ws1, 'S54', Ma_nt_value)
                try:
                    ws1['S54'].font = Font(name='Courier New', size=8)
                except Exception:
                    pass
            
            if 'TEN_HQ' in dtokhaimd_row:
                safe_set_cell(ws1, 'J7', dtokhaimd_row['TEN_HQ'])
                safe_set_cell(ws2, 'J7', dtokhaimd_row['TEN_HQ'])
            # Xử lý F8 = NGAY_DK + GIO_DK
            if 'NGAY_DK' in dtokhaimd_row and 'GIO_DK' in vnaccs_row:
                print(f"Debug F8 - NGAY_DK: {dtokhaimd_row['NGAY_DK']}, GIO_DK: {vnaccs_row['GIO_DK']}")
                f8_value = format_datetime_direct(dtokhaimd_row['NGAY_DK'], vnaccs_row['GIO_DK'])
                safe_set_cell(ws1, 'F8', f8_value)
                safe_set_cell(ws2, 'F8', f8_value)
            elif 'NGAY_DK' in dtokhaimd_row:
                safe_set_cell(ws1, 'F8', format_date(dtokhaimd_row['NGAY_DK']))
                safe_set_cell(ws2, 'F8', format_date(dtokhaimd_row['NGAY_DK']))
            if 'MA_DV' in dtokhaimd_row:
                safe_set_cell(ws1, 'F13', dtokhaimd_row['MA_DV'])
            
            if 'MA_BC_DV' in dtokhaimd_row:
                safe_set_cell(ws1, 'F16', dtokhaimd_row['MA_BC_DV'])
            
            if 'DIA_CHI_DV' in dtokhaimd_row:
                safe_set_cell(ws1, 'F17', dtokhaimd_row['DIA_CHI_DV'])
            
            if 'SO_DT_DV' in dtokhaimd_row:
                safe_set_cell(ws1, 'F19', dtokhaimd_row['SO_DT_DV'])
            
            if 'DV_DT' in dtokhaimd_row:
                safe_set_cell(ws1, 'F30', dtokhaimd_row['DV_DT'])
            
            if 'TEN_PTVT' in dtokhaimd_row:
                safe_set_cell(ws1, 'M45', dtokhaimd_row['TEN_PTVT'])
            
            if 'NGAYKH' in dtokhaimd_row:
                safe_set_cell(ws1, 'I46', format_date(dtokhaimd_row['NGAYKH']))
            
            if 'NGAYDEN' in dtokhaimd_row:
                safe_set_cell(ws1, 'M78', format_date(dtokhaimd_row['NGAYDEN']))
            
            if 'VAN_DON' in dtokhaimd_row:
                safe_set_cell(ws1, 'H39', dtokhaimd_row['VAN_DON'])
            
            if 'MA_CK' in dtokhaimd_row:
                safe_set_cell(ws1, 'I43', dtokhaimd_row['MA_CK'])
            
            if 'TEN_CK' in dtokhaimd_row:
                safe_set_cell(ws1, 'M44', dtokhaimd_row['TEN_CK'])
            
            if 'MA_CANGNN' in dtokhaimd_row:
                safe_set_cell(ws1, 'I44', dtokhaimd_row['MA_CANGNN'])
            
            if 'CANGNN' in dtokhaimd_row:
                safe_set_cell(ws1, 'M43', dtokhaimd_row['CANGNN'])
            
            if 'NUOC_NK' in dtokhaimd_row:
                safe_set_cell(ws1, 'F35', dtokhaimd_row['NUOC_NK'])
            magh_mant_value = f"{dtokhaimd_row['MA_GH']} - {dtokhaimd_row['MA_NT']}"
            if 'MA_GH' in dtokhaimd_row:
                safe_set_cell(ws1, 'Q53', magh_mant_value)
            
            if 'SOHANG' in dtokhaimd_row:
                safe_set_cell(ws1, 'Z61', format_number_vn(dtokhaimd_row['SOHANG']))
            
            if 'MA_PTTT' in dtokhaimd_row:
                safe_set_cell(ws1, 'S52', dtokhaimd_row['MA_PTTT'])
            
            # if 'MA_NT' in dtokhaimd_row:
            #     safe_set_cell(ws1, 'Q53', dtokhaimd_row['MA_NT'])
            
            if 'TYGIA_VND' in dtokhaimd_row:
                safe_set_cell(ws1, 'S55', format_number_vn(dtokhaimd_row['TYGIA_VND']))
            
            # Bổ sung các trường còn thiếu
            if 'TONGTGKB' in dtokhaimd_row:
                safe_set_cell(ws1, 'U53', format_number_vn(dtokhaimd_row['TONGTGKB']))
            
            if 'TONGTGTT' in dtokhaimd_row:
                safe_set_cell(ws1, 'U54', format_number_vn(dtokhaimd_row['TONGTGTT']))
            
            
            
            if 'TR_LUONG' in dtokhaimd_row:
                safe_set_cell(ws1, 'H41', format_number_vn(dtokhaimd_row['TR_LUONG']))
            
            if 'DVT_TR_LUONG' in dtokhaimd_row:
                safe_set_cell(ws1, 'M41', dtokhaimd_row['DVT_TR_LUONG'])
            
            if 'SO_KIEN' in dtokhaimd_row:
                safe_set_cell(ws1, 'H40', format_number_vn(dtokhaimd_row['SO_KIEN']))
            
            if 'DVT_KIEN' in dtokhaimd_row:
                safe_set_cell(ws1, 'M40', dtokhaimd_row['DVT_KIEN'])
            
            if 'KY_HIEU_SO_HIEU' in dtokhaimd_row:
                safe_set_cell(ws1, 'H47', dtokhaimd_row['KY_HIEU_SO_HIEU'])
            
            if 'MA_HDTM' in dtokhaimd_row:
                safe_set_cell(ws1, 'P49', dtokhaimd_row['MA_HDTM'])
            
            if 'SO_HDTM' in dtokhaimd_row:
                safe_set_cell(ws1, 'R49', dtokhaimd_row['SO_HDTM'])
            
            if 'NGAY_HDTM' in dtokhaimd_row:
                safe_set_cell(ws1, 'S51', format_date(dtokhaimd_row['NGAY_HDTM']))
            
            if 'MA_PL_GIA_HDTM' in dtokhaimd_row:
                safe_set_cell(ws1, 'AA53', dtokhaimd_row['MA_PL_GIA_HDTM'])
            
            if 'TONGTG_HDTM' in dtokhaimd_row:
                safe_set_cell(ws1, 'U53', format_number_vn(dtokhaimd_row['TONGTG_HDTM']))
            
            if 'THUE' in dtokhaimd_row:
                safe_set_cell(ws1, 'N57', format_number_vn(dtokhaimd_row['THUE']))
            
            if '_Ten_DV_L1' in dtokhaimd_row:
                safe_set_cell(ws1, 'F14', dtokhaimd_row['_Ten_DV_L1'])
            
            if 'SoHSTK' in dtokhaimd_row:
                safe_set_cell(ws1, 'I66', dtokhaimd_row['SoHSTK'])
            
            if 'MA_THOI_HAN_NOP_THUE' in dtokhaimd_row:
                safe_set_cell(ws1, 'V57', dtokhaimd_row['MA_THOI_HAN_NOP_THUE'])
        
        # Từ bảng DTOKHAIMD_VNACCS
        if vnaccs_row is not None:
            if 'TRUNG_CHUYEN_DIEM_CUOI' in vnaccs_row:
                safe_set_cell(ws1, 'K78', vnaccs_row['TRUNG_CHUYEN_DIEM_CUOI'])
            
            if 'TRUNG_CHUYEN_NGAY_KT' in vnaccs_row:
                safe_set_cell(ws1, 'M73', format_date(vnaccs_row['TRUNG_CHUYEN_NGAY_KT']))
            
            if 'TRUNG_CHUYEN_GHI_CHU' in vnaccs_row:
                safe_set_cell(ws1, 'F64', vnaccs_row['TRUNG_CHUYEN_GHI_CHU'])
            e95_value = ""
            if 'MA_DIEM_XEP_HANG_LEN_XE' in vnaccs_row:
                e95_value = f"1  {vnaccs2_row['MA_DD_LUU_KHO']}  2    3    4    5"
                safe_set_cell(ws1, 'E95', e95_value)
            
            if 'TEN_DIEM_XEP_HANG_LEN_XE' in vnaccs_row:
                safe_set_cell(ws1, 'E96', vnaccs_row['TEN_DIEM_XEP_HANG_LEN_XE'])
            
            if 'DIA_CHI_DIEM_XEP_HANG_LEN_XE' in vnaccs_row:
                safe_set_cell(ws1, 'E97', vnaccs_row['DIA_CHI_DIEM_XEP_HANG_LEN_XE'])
            
            if 'MA_PL_KTRA' in vnaccs_row:
                safe_set_cell(ws1, 'F6', vnaccs_row['MA_PL_KTRA'])
                safe_set_cell(ws2, 'F6', vnaccs_row['MA_PL_KTRA'])
            if 'MA_HANG_HOA_DD' in vnaccs_row:
                safe_set_cell(ws1, 'Y6', vnaccs_row['MA_HANG_HOA_DD'])
                safe_set_cell(ws2, 'Y6', vnaccs_row['MA_HANG_HOA_DD'])
            if 'TEN_DD_LUU_KHO' in vnaccs_row:
                safe_set_cell(ws1, 'M42', vnaccs_row['TEN_DD_LUU_KHO'])
            
            if 'MA_PL_NOP_THUE' in vnaccs_row:
                safe_set_cell(ws1, 'Z57', vnaccs_row['MA_PL_NOP_THUE'])
            
            if 'SO_TRANG_TK' in vnaccs_row:
                safe_set_cell(ws1, 'T61', format_number_vn(vnaccs_row['SO_TRANG_TK']))
            
            if 'MA_QLY_USER' in vnaccs_row:
                safe_set_cell(ws1, 'X66', vnaccs_row['MA_QLY_USER'])
            
            if 'TEN_TRUONG_DV_HQ' in vnaccs_row:
                safe_set_cell(ws1, 'I70', vnaccs_row['TEN_TRUONG_DV_HQ'])
            
            # Xử lý I71 = NGAY_HOAN_THANH_KT + GIO_HOAN_THANH_KT
            if 'NGAY_HOAN_THANH_KT' in vnaccs_row and 'GIO_HOAN_THANH_KT' in vnaccs_row:
                print(f"Debug I71 - NGAY_HOAN_THANH_KT: {vnaccs_row['NGAY_HOAN_THANH_KT']}, GIO_HOAN_THANH_KT: {vnaccs_row['GIO_HOAN_THANH_KT']}")
                i71_value = format_datetime_direct(vnaccs_row['NGAY_HOAN_THANH_KT'], vnaccs_row['GIO_HOAN_THANH_KT'])
                safe_set_cell(ws1, 'I71', i71_value)
            elif 'NGAY_HOAN_THANH_KT' in vnaccs_row:
                safe_set_cell(ws1, 'I71', format_date(vnaccs_row['NGAY_HOAN_THANH_KT']))
            
            # Xử lý I72 = NGAY_CAP_PHEP + GIO_CAP_PHEP
            if 'NGAY_CAP_PHEP' in vnaccs_row and 'GIO_CAP_PHEP' in vnaccs_row:
                print(f"Debug I72 - NGAY_CAP_PHEP: {vnaccs_row['NGAY_CAP_PHEP']}, GIO_CAP_PHEP: {vnaccs_row['GIO_CAP_PHEP']}")
                i72_value = format_datetime_direct(vnaccs_row['NGAY_CAP_PHEP'], vnaccs_row['GIO_CAP_PHEP'])
                safe_set_cell(ws1, 'I72', i72_value)
            elif 'NGAY_CAP_PHEP' in vnaccs_row:
                safe_set_cell(ws1, 'I72', format_date(vnaccs_row['NGAY_CAP_PHEP']))
            if 'NGAY_DK_SUA' in vnaccs_row and 'GIO_DK_SUA' in vnaccs_row:
                print(f"Debug I71 - NGAY_HOAN_THANH_KT: {vnaccs_row['NGAY_HOAN_THANH_KT']}, GIO_HOAN_THANH_KT: {vnaccs_row['GIO_HOAN_THANH_KT']}")
                Y8_value = format_datetime_direct(vnaccs_row['NGAY_DK_SUA'], vnaccs_row['GIO_DK_SUA'])
                safe_set_cell(ws1, 'Y8', Y8_value)
                safe_set_cell(ws2, 'Y8', Y8_value)
            elif 'NGAY_DK_SUA' in vnaccs_row:
                safe_set_cell(ws1, 'Y8', format_date(vnaccs_row['NGAY_DK_SUA']))
                safe_set_cell(ws2, 'Y8', format_date(vnaccs_row['NGAY_DK_SUA']))
        # Từ bảng DTOKHAIMD_VNACCS2
        if vnaccs2_row is not None:
            if 'MA_HIEU_PTVC' in vnaccs2_row:
                safe_set_cell(ws1, 'L6', vnaccs2_row['MA_HIEU_PTVC'])
            
            if 'NHOM_HO_SO' in vnaccs2_row:
                safe_set_cell(ws1, 'Y7', vnaccs2_row['NHOM_HO_SO'])
                safe_set_cell(ws2, 'Y7', vnaccs2_row['NHOM_HO_SO'])
            
            if 'DIA_CHI_DT1' in vnaccs2_row:
                safe_set_cell(ws1, 'F33', vnaccs2_row['DIA_CHI_DT1'])
            
            if 'DIA_CHI_DT2' in vnaccs2_row:
                safe_set_cell(ws1, 'R33', vnaccs2_row['DIA_CHI_DT2'])
            
            if 'DIA_CHI_DT3' in vnaccs2_row:
                safe_set_cell(ws1, 'F34', vnaccs2_row['DIA_CHI_DT3'])
            
            if 'DIA_CHI_DT4' in vnaccs2_row:
                safe_set_cell(ws1, 'R34', vnaccs2_row['DIA_CHI_DT4'])
            
            if 'MA_DD_LUU_KHO' in vnaccs2_row:
                safe_set_cell(ws1, 'I42', vnaccs2_row['MA_DD_LUU_KHO'])
            
            if 'CHI_THI_HQ_NGAY' in vnaccs2_row:
                chi_thi_ngay = vnaccs2_row.get('CHI_THI_HQ_NGAY')
                safe_set_cell(ws1, 'D90', format_date(chi_thi_ngay) if chi_thi_ngay else '')
                chi_thi_hq_ten = vnaccs2_row.get('CHI_THI_HQ_TEN') or ''
                safe_set_cell(ws1, 'I90', str(chi_thi_hq_ten).strip())
                chi_thi_hq_nd = vnaccs2_row.get('CHI_THI_HQ_ND') or ''
                safe_set_cell(ws1, 'R90', str(chi_thi_hq_nd).strip())
        ws1.sheet_properties.pageSetUpPr.fitToPage = True
        ws1.page_setup.fitToWidth = 1
        ws1.page_setup.fitToHeight = 0
        ws1.page_setup.scale = None
        
        # Xóa tiêu đề lặp lại theo trang (nếu có)
        try:
            ws1.print_title_rows = None
            ws1.print_title_cols = None
        except Exception:
            pass
        
        # Xóa toàn bộ ngắt trang cũ trước khi thiết lập lại
        try:
            from openpyxl.worksheet.pagebreak import PageBreak
            ws1.row_breaks = PageBreak()
            ws1.col_breaks = PageBreak()
        except Exception:
            pass
        
        goods_df = data.get('dhangmddk', pd.DataFrame())
        if not isinstance(goods_df, pd.DataFrame):
            goods_df = pd.DataFrame(goods_df or [])
        if not goods_df.empty:
            goods_df = goods_df.reset_index(drop=True)
        goods_row_count = len(goods_df)

        declared_goods_rows = 0
        if dtokhaimd_row is not None and 'SOHANG' in dtokhaimd_row:
            try:
                declared_goods_rows = int(dtokhaimd_row['SOHANG'])
            except (ValueError, TypeError):
                declared_goods_rows = 0

        goods_page_count = math.ceil(goods_row_count / 2) if goods_row_count else 0
        declared_page_count = math.ceil(declared_goods_rows / 2) if declared_goods_rows else 0
        so_trang_phu = max(goods_page_count, declared_page_count)

        copied_pages = 0
        last_page_partial = False
        line_in_page = 0

        def reset_tkx_detail_page():
            clear_cells_by_mapping(ws2, TKX_DETAIL_ODD_CLEAR_CELLS, safe_set_cell)
            clear_cells_by_mapping(ws2, TKX_DETAIL_EVEN_CLEAR_CELLS, safe_set_cell)

        def copy_current_page(partial=False):
            nonlocal copied_pages, last_page_partial
            start_row = TKX_DETAIL_FIRST_DETAIL_ROW + copied_pages * TKX_DETAIL_PAGE_HEIGHT
            cell_range = TKX_DETAIL_TEMPLATE_HALF_RANGE if partial else TKX_DETAIL_TEMPLATE_FULL_RANGE
            copy_row_range_with_style(ws2, ws1, cell_range, start_row)
            last_page_partial = partial
            copied_pages += 1

        if goods_row_count:
            reset_tkx_detail_page()
            for idx in range(goods_row_count):
                if line_in_page == 0 and idx != 0:
                    reset_tkx_detail_page()
                goods_row = goods_df.iloc[idx]
                current_mapping = TKX_DETAIL_ODD_FIELD_MAP if line_in_page == 0 else TKX_DETAIL_EVEN_FIELD_MAP
                populate_worksheet_from_mapping(ws2, goods_row, current_mapping, safe_set_cell)
                line_in_page += 1
                if line_in_page == 2:
                    copy_current_page(partial=False)
                    line_in_page = 0
            if line_in_page == 1:
                copy_current_page(partial=True)
                line_in_page = 0
        else:
            reset_tkx_detail_page()

        while copied_pages < so_trang_phu:
            reset_tkx_detail_page()
            copy_current_page(partial=False)

        so_trang_phu = copied_pages

        # Thiết lập khu vực in theo tổng số hàng (header 138 hàng + n*53)
        if so_trang_phu == 0:
            tong_hang = TKX_DETAIL_HEADER_ROWS
        else:
            body_height = (so_trang_phu - 1) * TKX_DETAIL_PAGE_HEIGHT
            last_height = TKX_DETAIL_PARTIAL_HEIGHT if last_page_partial else TKX_DETAIL_PAGE_HEIGHT
            tong_hang = TKX_DETAIL_HEADER_ROWS + body_height + last_height
        ws1.print_area = f"A1:AC{tong_hang}"
        ws1.print_options.horizontalCentered = True
        
        # Thiết lập ngắt trang: trang 1 tại 75, trang 2 tại 138; các trang sau: 138 + i*53
        try:
            break_positions = []
            ws1.row_breaks.append(Break(id=80))
            break_positions.append(80)
            ws1.row_breaks.append(Break(id=144))
            break_positions.append(144)
            for i in range(1, so_trang_phu + 1):
                if last_page_partial and i == so_trang_phu:
                    r = TKX_DETAIL_HEADER_ROWS + (i - 1) * TKX_DETAIL_PAGE_HEIGHT + TKX_DETAIL_PARTIAL_HEIGHT
                else:
                    r = TKX_DETAIL_HEADER_ROWS + i * TKX_DETAIL_PAGE_HEIGHT
                ws1.row_breaks.append(Break(id=r))
                break_positions.append(r)
            if break_positions:
                print(f"Manual page breaks at rows: {break_positions}")
        except Exception:
            pass
        if ws2.title in wb.sheetnames:
            wb.remove(ws2)
        try:
            ws1.title = 'TKX'
        except Exception:
            pass
        # Lưu file
        wb.save(output_file)
        print(f"✓ Đã tạo file TKX: {output_file}")
        
    except Exception as e:
        print(f"✗ Lỗi khi điền dữ liệu vào Excel: {e}")
        messagebox.showerror('Excel Error', f'Lỗi khi điền dữ liệu: {e}')

def format_datetime_vn(date_str, time_str):
    """
    Định dạng ngày giờ theo kiểu Việt Nam
    Ví dụ: date_str='04-09-2025', time_str='105059' -> '04/09/2025 10:50:59'
    """
    if not date_str or not time_str:
        return date_str
    
    try:
        # Xử lý date_str - có thể là string hoặc đã được format_date xử lý
        if isinstance(date_str, str):
            # Nếu là string, parse theo format dd-mm-yyyy
            if '-' in date_str:
                day, month, year = date_str.split('-')
            elif '/' in date_str:
                # Nếu đã được format_date xử lý thành dd/mm/yyyy
                day, month, year = date_str.split('/')
            else:
                return date_str
        else:
            # Nếu là datetime object, convert về string
            date_str = str(date_str)
            if '-' in date_str:
                day, month, year = date_str.split('-')
            elif '/' in date_str:
                day, month, year = date_str.split('/')
            else:
                return date_str
        
        # Parse giờ từ format hhmmss
        if len(str(time_str)) == 6:
            time_str = str(time_str)
            hour = time_str[:2]
            minute = time_str[2:4]
            second = time_str[4:6]
        else:
            return date_str
        
        # Tạo datetime string theo format dd/mm/yyyy hh:mm:ss
        return f"{day}/{month}/{year} {hour}:{minute}:{second}"
        
    except Exception as e:
        print(f"Error in format_datetime_vn: {e}")
        return date_str

def format_datetime_direct(date_str, time_str):
    """
    Định dạng ngày giờ trực tiếp từ database
    Ví dụ: date_str='2025-09-04 00:00:00', time_str=105059.0 -> '04/09/2025 10:50:59'
    """
    if not date_str or not time_str:
        return date_str
    
    try:
        # Xử lý date_str - có thể là pandas.Timestamp
        if hasattr(date_str, 'strftime'):
            # Nếu là pandas.Timestamp, lấy ngày theo format yyyy-mm-dd
            date_str = date_str.strftime('%Y-%m-%d')
        
        # Chuyển về string
        date_str = str(date_str).strip()
        
        # Parse ngày từ format yyyy-mm-dd
        if '-' in date_str:
            parts = date_str.split('-')
            if len(parts) == 3:
                year, month, day = parts
            else:
                return date_str
        else:
            return date_str
        
        # Xử lý time_str - có thể là numpy.float64
        time_str = str(int(float(time_str))).zfill(6)  # Chuyển về int và đảm bảo 6 chữ số
        
        # Parse giờ từ format hhmmss
        if len(time_str) == 6:
            hour = time_str[:2]
            minute = time_str[2:4]
            second = time_str[4:6]
        else:
            return date_str
        
        # Tạo datetime string theo format dd/mm/yyyy hh:mm:ss
        result = f"{day}/{month}/{year} {hour}:{minute}:{second}"
        return result
        
    except Exception as e:
        print(f"Error in format_datetime_direct: {e}")
        return date_str

def PrintCD_TKN(dtokhaimdid, index, data, output_path):
    """In/ghi Tờ khai xuất (TKX) dựa vào template và dữ liệu"""
    os.makedirs(output_path, exist_ok=True)
    template_path = 'Excel_Tempt/TKN_TEMPT.xlsx'
    wb = load_workbook(_get_template_stream(template_path))
    ws1 = wb['TKN_Tempt1']
    ws2 = wb['TKN_Tempt2']
    
    # Lấy dữ liệu từ các bảng
    dtokhaimd_row = None
    vnaccs_row = None
    vnaccs2_row = None
    
    if not data['dtokhaimd'].empty:
        dtokhaimd_row = data['dtokhaimd'].iloc[0]
    if not data['vnaccs'].empty:
        vnaccs_row = data['vnaccs'].iloc[0]
    if not data['vnaccs2'].empty:
        vnaccs2_row = data['vnaccs2'].iloc[0]
    
    sotk_value = ''
    if dtokhaimd_row is not None and 'SOTK' in dtokhaimd_row:
        sotk_value = _format_sotk_value(dtokhaimd_row['SOTK'])
    if not sotk_value:
        sotk_value = _format_sotk_value(dtokhaimdid)
    output_file = os.path.join(output_path, f'ToKhaiHQ7N_QDTQ_{sotk_value}.xlsx')
    
    total_pages_declared = None
    if dtokhaimd_row is not None and 'SO_TRANG_TK' in dtokhaimd_row:
        try:
            total_pages_declared = int(dtokhaimd_row['SO_TRANG_TK'])
        except (ValueError, TypeError):
            total_pages_declared = None
    
    def safe_set_cell(worksheet, cell_address, value):
        """Ghi giá trị vào ô một cách an toàn, xử lý merged cells"""
        try:
            cell = worksheet[cell_address]
            if isinstance(cell, MergedCell):
                target = None
                for merged_range in worksheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        target = worksheet.cell(merged_range.min_row, merged_range.min_col)
                        break
                if target is None:
                    row = getattr(cell, 'row', None)
                    col = getattr(cell, 'col_idx', None) or getattr(cell, 'column', None)
                    if row is not None and col is not None:
                        target = worksheet.cell(row=row, column=col)
                (target or cell).value = value
            else:
                cell.value = value
        except Exception as e:
            print(f"Warning: Không thể ghi vào ô {cell_address}: {e}")
    
    # Điền dữ liệu vào ws1 theo mapping
    try:
        # Từ bảng DTOKHAIMD
        if dtokhaimd_row is not None:
            # Thông tin cơ bản
            mvvalue = ""
            if 'SOTK' in dtokhaimd_row:
                mvvalue=f"*{dtokhaimd_row['SOTK']}*"
                safe_set_cell(ws1, 'AA3', mvvalue)
                safe_set_cell(ws1, 'E4', dtokhaimd_row['SOTK'])
                safe_set_cell(ws1, 'E79', dtokhaimd_row['SOTK'])
                safe_set_cell(ws2, 'E4', dtokhaimd_row['SOTK'])
            #if 'MA_LH' in dtokhaimd_row:
            if dtokhaimd_row['TTTK'] == 'P':
                trangthaitokhai = 'Tờ khai hàng hóa nhập khẩu (thông báo kết quả phân luồng)'
                safe_set_cell(ws1, 'F2', trangthaitokhai)
                safe_set_cell(ws2, 'D2', trangthaitokhai)    #safe_set_cell(ws1, 'P6', dtokhaimd_row['MA_LH'])
            elif dtokhaimd_row['TTTK'] == 'T':
                trangthaitokhai = 'Tờ khai hàng hóa nhập khẩu (thông quan)'
                safe_set_cell(ws1, 'F2', trangthaitokhai)
                safe_set_cell(ws2, 'D2', trangthaitokhai)
            elif dtokhaimd_row['TTTK'] == 'E':
                trangthaitokhai = 'Bản xác nhận nội dung tờ khai hàng hóa nhập khẩu<In thử>'
                safe_set_cell(ws1, 'F2', trangthaitokhai)
                safe_set_cell(ws2, 'D2', trangthaitokhai)
            if 'TEN_HQ' in dtokhaimd_row:
                safe_set_cell(ws1, 'L7', dtokhaimd_row['TEN_HQ'])
                safe_set_cell(ws1, 'L82', dtokhaimd_row['TEN_HQ'])
                safe_set_cell(ws2, 'L7', dtokhaimd_row['TEN_HQ'])
            
            # Xử lý G8 = NGAY_DK + GIO_DK
            if 'NGAY_DK' in dtokhaimd_row and 'GIO_DK' in vnaccs_row:
                print(f"Debug G8 - NGAY_DK: {dtokhaimd_row['NGAY_DK']}, GIO_DK: {vnaccs_row['GIO_DK']}")
                g8_value = format_datetime_direct(dtokhaimd_row['NGAY_DK'], vnaccs_row['GIO_DK'])
                safe_set_cell(ws1, 'G8', g8_value)
                safe_set_cell(ws1, 'G83', g8_value)
                safe_set_cell(ws2, 'G8', g8_value)
            elif 'NGAY_DK' in dtokhaimd_row:
                safe_set_cell(ws1, 'G8', format_date(dtokhaimd_row['NGAY_DK']))
                safe_set_cell(ws2, 'G8', format_date(dtokhaimd_row['NGAY_DK']))
            
            if 'MA_DV' in dtokhaimd_row:
                safe_set_cell(ws1, 'H10', dtokhaimd_row['MA_DV'])
            if 'TRUNG_CHUYEN_GHI_CHU' in vnaccs_row:
                safe_set_cell(ws1, 'G85', vnaccs_row['TRUNG_CHUYEN_GHI_CHU'])
            if 'MA_BC_DV' in dtokhaimd_row:
                safe_set_cell(ws1, 'H13', dtokhaimd_row['MA_BC_DV'])
            
            if 'DIA_CHI_DV' in dtokhaimd_row:
                safe_set_cell(ws1, 'H14', dtokhaimd_row['DIA_CHI_DV'])
            
            if 'SO_DT_DV' in dtokhaimd_row:
                safe_set_cell(ws1, 'H16', dtokhaimd_row['SO_DT_DV'])
            
            if 'DV_DT' in dtokhaimd_row:
                safe_set_cell(ws1, 'H23', dtokhaimd_row['DV_DT'])
            
            if 'MA_BC_DT' in dtokhaimd_row:
                safe_set_cell(ws1, 'H24', dtokhaimd_row['MA_BC_DT'])
            
            if 'MA_PTVT' in dtokhaimd_row:
                safe_set_cell(ws1, 'T34', dtokhaimd_row['MA_PTVT'])
            
            if 'TEN_PTVT' in dtokhaimd_row:
                safe_set_cell(ws1, 'Z34', dtokhaimd_row['TEN_PTVT'])
            
            if 'NGAYDEN' in dtokhaimd_row:
                safe_set_cell(ws1, 'U35', format_date(dtokhaimd_row['NGAYDEN']))
            
            if 'VAN_DON' in dtokhaimd_row:
                safe_set_cell(ws1, 'D31', dtokhaimd_row['VAN_DON'])
            
            if 'MA_CK' in dtokhaimd_row:
                safe_set_cell(ws1, 'U31', dtokhaimd_row['MA_CK'])
            
            if 'TEN_CK' in dtokhaimd_row:
                safe_set_cell(ws1, 'Z31', dtokhaimd_row['TEN_CK'])
            
            if 'MA_CANGNN' in dtokhaimd_row:
                safe_set_cell(ws1, 'U32', dtokhaimd_row['MA_CANGNN'])
            
            if 'CANGNN' in dtokhaimd_row:
                safe_set_cell(ws1, 'Z32', dtokhaimd_row['CANGNN'])
            
            if 'MA_GP' in dtokhaimd_row:
                safe_set_cell(ws1, 'D50', dtokhaimd_row['MA_GP'])
            
            if 'SO_GP' in dtokhaimd_row:
                safe_set_cell(ws1, 'F50', dtokhaimd_row['SO_GP'])
            
            if 'NUOC_XK' in dtokhaimd_row:
                safe_set_cell(ws1, 'H27', dtokhaimd_row['NUOC_XK'])
            
            #if 'MA_GH' in dtokhaimd_row:
                #safe_set_cell(ws1, 'J45', dtokhaimd_row['MA_GH'])
            
            if 'SOHANG' in dtokhaimd_row:
                safe_set_cell(ws1, 'AF75', format_number_vn(dtokhaimd_row['SOHANG']))
            
            if 'MA_PTTT' in dtokhaimd_row:
                safe_set_cell(ws1, 'J44', dtokhaimd_row['MA_PTTT'])
            
            #if 'MA_NT' in dtokhaimd_row:
                #safe_set_cell(ws1, 'J45', dtokhaimd_row['MA_NT'])
            
            if 'TYGIA_VND' in dtokhaimd_row:
                safe_set_cell(ws1, 'AB70', format_number_vn(dtokhaimd_row['TYGIA_VND']))
            
            if 'MA_NT_TY_GIA_VND' in dtokhaimd_row:
                safe_set_cell(ws1, 'X70', dtokhaimd_row['MA_NT_TY_GIA_VND'])
            
            ma_nt_phi_bh_key = None
            if 'MA_NT_PHI_BH' in dtokhaimd_row:
                ma_nt_phi_bh_key = 'MA_NT_PHI_BH'
            elif 'MA_NT_PHI_BH ' in dtokhaimd_row:
                ma_nt_phi_bh_key = 'MA_NT_PHI_BH '
            
            if 'MA_PHI_BH' in dtokhaimd_row:
                ma_phi_bh_val = dtokhaimd_row['MA_PHI_BH']
                ma_nt_phi_bh_val = dtokhaimd_row.get(ma_nt_phi_bh_key) if ma_nt_phi_bh_key else None
                ma_phi_bh = str(ma_phi_bh_val).strip() if ma_phi_bh_val is not None else ''
                ma_nt_phi_bh = str(ma_nt_phi_bh_val).strip() if ma_nt_phi_bh_val is not None else ''
                safe_set_cell(ws1, 'I56', f"{ma_phi_bh} - {ma_nt_phi_bh} - ")
                if ma_nt_phi_bh and 'PHI_BH' in dtokhaimd_row and dtokhaimd_row['PHI_BH'] is not None:
                    safe_set_cell(ws1, 'L56', format_number_vn(dtokhaimd_row['PHI_BH']))
                else:
                    safe_set_cell(ws1, 'L56', "")
            
            ma_nt_phi_vc_key = None
            if 'MA_NT_PHI_VC' in dtokhaimd_row:
                ma_nt_phi_vc_key = 'MA_NT_PHI_VC'
            elif 'MA_NT_PHI_VC ' in dtokhaimd_row:
                ma_nt_phi_vc_key = 'MA_NT_PHI_VC '
            
            if 'MA_PHI_VC' in dtokhaimd_row and ma_nt_phi_vc_key:
                ma_phi_vc_val = dtokhaimd_row['MA_PHI_VC']
                ma_nt_phi_vc_val = dtokhaimd_row[ma_nt_phi_vc_key]
                ma_phi_vc = str(ma_phi_vc_val).strip() if ma_phi_vc_val is not None else ''
                ma_nt_phi_vc = str(ma_nt_phi_vc_val).strip() if ma_nt_phi_vc_val is not None else ''
                safe_set_cell(ws1, 'I55', f"{ma_phi_vc} - {ma_nt_phi_vc} - ")
            elif 'MA_PHI_VC' in dtokhaimd_row:
                safe_set_cell(ws1, 'I55', dtokhaimd_row['MA_PHI_VC'])
            
            if 'PHI_VC' in dtokhaimd_row:
                safe_set_cell(ws1, 'L55', format_number_vn(dtokhaimd_row['PHI_VC']))
            
            if 'TONGTGKB' in dtokhaimd_row:
                safe_set_cell(ws1, 'P45', format_number_vn(dtokhaimd_row['TONGTGKB']))
            
            if 'TONGTGTT' in dtokhaimd_row:
                safe_set_cell(ws1, 'J46', format_number_vn(dtokhaimd_row['TONGTGTT']))
            
            if 'TR_LUONG' in dtokhaimd_row:
                safe_set_cell(ws1, 'K37', format_number_vn(dtokhaimd_row['TR_LUONG']))
            
            if 'DVT_TR_LUONG' in dtokhaimd_row:
                safe_set_cell(ws1, 'P37', dtokhaimd_row['DVT_TR_LUONG'])
            
            if 'SO_KIEN' in dtokhaimd_row:
                safe_set_cell(ws1, 'K36', format_number_vn(dtokhaimd_row['SO_KIEN']))
            
            if 'DVT_KIEN' in dtokhaimd_row:
                safe_set_cell(ws1, 'P36', dtokhaimd_row['DVT_KIEN'])
            
            if 'SO_CONTAINER' in dtokhaimd_row:
                safe_set_cell(ws1, 'K38', format_number_vn(dtokhaimd_row['SO_CONTAINER']))
            
            #if 'MA_HDTM' in dtokhaimd_row:
                #safe_set_cell(ws1, 'J41', dtokhaimd_row['MA_HDTM'])
            
            #if 'SO_HDTM' in dtokhaimd_row:
                #safe_set_cell(ws1, 'J41', dtokhaimd_row['SO_HDTM'])
            
            if 'NGAY_HDTM' in dtokhaimd_row:
                safe_set_cell(ws1, 'J43', format_date(dtokhaimd_row['NGAY_HDTM']))
            
            #if 'MA_PL_GIA_HDTM' in dtokhaimd_row:
                #safe_set_cell(ws1, 'J45', dtokhaimd_row['MA_PL_GIA_HDTM'])
            
            if 'TONGTG_HDTM' in dtokhaimd_row:
                safe_set_cell(ws1, 'P45', format_number_vn(dtokhaimd_row['TONGTG_HDTM']))
            
            #if 'THUE' in dtokhaimd_row:
                #safe_set_cell(ws1, 'L55', format_number_vn(dtokhaimd_row['THUE']))
            
            if '_Ten_DV_L1' in dtokhaimd_row:
                safe_set_cell(ws1, 'H11', dtokhaimd_row['_Ten_DV_L1'])
            
            if 'TongTienThue' in dtokhaimd_row:
                safe_set_cell(ws1, 'X68', format_number_vn(dtokhaimd_row['TongTienThue']))
            
            if 'MA_THOI_HAN_NOP_THUE' in dtokhaimd_row:
                safe_set_cell(ws1, 'X73', dtokhaimd_row['MA_THOI_HAN_NOP_THUE'])
        
        # Từ bảng DTOKHAIMD_VNACCS
        if vnaccs_row is not None:
            def write_tax_code_name(code_key, name_key, cell_address):
                code = str(vnaccs_row.get(code_key, '') or '').strip()
                name = str(vnaccs_row.get(name_key, '') or '').strip()
                if code and name:
                    value = f'{code}  {name}'
                else:
                    value = code or name
                if value:
                    safe_set_cell(ws1, cell_address, value)

            def write_tax_amount(amount_key, cell_address):
                value = vnaccs_row.get(amount_key, None)
                if value not in (None, ''):
                    safe_set_cell(ws1, cell_address, format_number_vn(value))

            if 'MA_KHAI_TRGIA' in vnaccs_row:
                safe_set_cell(ws1, 'I52', vnaccs_row['MA_KHAI_TRGIA'])
            
            if 'NOI_DUNG_KHAI_TRGIA' in vnaccs_row:
                safe_set_cell(ws1, 'D64', vnaccs_row['NOI_DUNG_KHAI_TRGIA'])
            
            if 'HS_PB_TRGIA' in vnaccs_row:
                safe_set_cell(ws1, 'J47', format_number_vn(vnaccs_row['HS_PB_TRGIA']))
            
            if 'MA_NGUOI_NOP_THUE' in vnaccs_row:
                safe_set_cell(ws1, 'AF73', vnaccs_row['MA_NGUOI_NOP_THUE'])
            
            if 'MA_VB_PQUY' in vnaccs_row:
                safe_set_cell(ws1, 'Y40', vnaccs_row['MA_VB_PQUY'])
            
            if 'MA_VB_PQUY2' in vnaccs_row:
                safe_set_cell(ws1, 'Y40', vnaccs_row['MA_VB_PQUY2'])
            
            #if 'MA_LOAI_DINH_KEM' in vnaccs_row:
                #safe_set_cell(ws1, 'K84', vnaccs_row['MA_LOAI_DINH_KEM'])
            
            #if 'SO_DINH_KEM' in vnaccs_row:
                #safe_set_cell(ws1, 'K84', vnaccs_row['SO_DINH_KEM'])
            
            if 'MA_PL_KTRA' in vnaccs_row:
                safe_set_cell(ws1, 'I6', vnaccs_row['MA_PL_KTRA'])
                safe_set_cell(ws1, 'I81', vnaccs_row['MA_PL_KTRA'])
                safe_set_cell(ws2, 'I6', vnaccs_row['MA_PL_KTRA'])
            
            if 'MA_HANG_HOA_DD' in vnaccs_row:
                safe_set_cell(ws1, 'AE6', vnaccs_row['MA_HANG_HOA_DD'])
                safe_set_cell(ws1, 'AE81', vnaccs_row['MA_HANG_HOA_DD'])
                safe_set_cell(ws2, 'AE6', vnaccs_row['MA_HANG_HOA_DD'])
            
            if 'TEN_DD_LUU_KHO' in vnaccs_row:
                safe_set_cell(ws1, 'Z30', vnaccs_row['TEN_DD_LUU_KHO'])
            
            if 'MA_PL_NOP_THUE' in vnaccs_row:
                safe_set_cell(ws1, 'AF74', vnaccs_row['MA_PL_NOP_THUE'])
            
            if 'SO_TRANG_TK' in vnaccs_row:
                safe_set_cell(ws1, 'U75', format_number_vn(vnaccs_row['SO_TRANG_TK']))
            
            if 'MA_QLY_USER' in vnaccs_row:
                safe_set_cell(ws1, 'AC87', vnaccs_row['MA_QLY_USER'])
            
            write_tax_code_name('MA_SAC_THUE', 'TEN_SAC_THUE', 'D68')
            write_tax_amount('TONG_THUE', 'H68')
            
            if 'SO_DONG_TONG_THUE' in vnaccs_row:
                safe_set_cell(ws1, 'N68', format_number_vn(vnaccs_row['SO_DONG_TONG_THUE']))
            
            write_tax_code_name('MA_SAC_THUE2', 'TEN_SAC_THUE2', 'D69')
            write_tax_amount('TONG_THUE2', 'H69')
            
            if 'SO_DONG_TONG_THUE2' in vnaccs_row:
                safe_set_cell(ws1, 'N69', format_number_vn(vnaccs_row['SO_DONG_TONG_THUE2']))

            write_tax_code_name('MA_SAC_THUE3', 'TEN_SAC_THUE3', 'D70')
            write_tax_amount('TONG_THUE3', 'H70')
            if 'SO_DONG_TONG_THUE3' in vnaccs_row:
                safe_set_cell(ws1, 'N70', format_number_vn(vnaccs_row['SO_DONG_TONG_THUE3']))
            write_tax_code_name('MA_SAC_THUE4', 'TEN_SAC_THUE4', 'D71')
            write_tax_amount('TONG_THUE4', 'H71')
            if 'SO_DONG_TONG_THUE4' in vnaccs_row:
                safe_set_cell(ws1, 'N71', format_number_vn(vnaccs_row['SO_DONG_TONG_THUE4']))
            write_tax_code_name('MA_SAC_THUE5', 'TEN_SAC_THUE5', 'D72')
            write_tax_amount('TONG_THUE5', 'H72')
            if 'SO_DONG_TONG_THUE5' in vnaccs_row:
                safe_set_cell(ws1, 'N72', format_number_vn(vnaccs_row['SO_DONG_TONG_THUE5']))
            write_tax_code_name('MA_SAC_THUE6', 'TEN_SAC_THUE6', 'D73')
            write_tax_amount('TONG_THUE6', 'H73')
            if 'SO_DONG_TONG_THUE6' in vnaccs_row:
                safe_set_cell(ws1, 'N73', format_number_vn(vnaccs_row['SO_DONG_TONG_THUE6']))              
            if 'TEN_TRUONG_DV_HQ' in vnaccs_row:
                safe_set_cell(ws1, 'N121', vnaccs_row['TEN_TRUONG_DV_HQ'])
            
            # Xử lý N122 = NGAY_CAP_PHEP + GIO_CAP_PHEP
            if 'NGAY_CAP_PHEP' in vnaccs_row and 'GIO_CAP_PHEP' in vnaccs_row:
                print(f"Debug N122 - NGAY_CAP_PHEP: {vnaccs_row['NGAY_CAP_PHEP']}, GIO_CAP_PHEP: {vnaccs_row['GIO_CAP_PHEP']}")
                n122_value = format_datetime_direct(vnaccs_row['NGAY_CAP_PHEP'], vnaccs_row['GIO_CAP_PHEP'])
                safe_set_cell(ws1, 'N122', n122_value)
            elif 'NGAY_CAP_PHEP' in vnaccs_row:
                safe_set_cell(ws1, 'N122', format_date(vnaccs_row['NGAY_CAP_PHEP']))
            
            # Xử lý N123 = NGAY_HOAN_THANH_KT + GIO_HOAN_THANH_KT
            if 'NGAY_HOAN_THANH_KT' in vnaccs_row and 'GIO_HOAN_THANH_KT' in vnaccs_row:
                print(f"Debug N123 - NGAY_HOAN_THANH_KT: {vnaccs_row['NGAY_HOAN_THANH_KT']}, GIO_HOAN_THANH_KT: {vnaccs_row['GIO_HOAN_THANH_KT']}")
                n123_value = format_datetime_direct(vnaccs_row['NGAY_HOAN_THANH_KT'], vnaccs_row['GIO_HOAN_THANH_KT'])
                safe_set_cell(ws1, 'N123', n123_value)
            elif 'NGAY_HOAN_THANH_KT' in vnaccs_row:
                safe_set_cell(ws1, 'N123', format_date(vnaccs_row['NGAY_HOAN_THANH_KT']))
            if 'NGAY_DK_SUA' in vnaccs_row and 'GIO_DK_SUA' in vnaccs_row:
                #print(f"Debug I71 - NGAY_HOAN_THANH_KT: {vnaccs_row['NGAY_HOAN_THANH_KT']}, GIO_HOAN_THANH_KT: {vnaccs_row['GIO_HOAN_THANH_KT']}")
                R8_value = format_datetime_direct(vnaccs_row['NGAY_DK_SUA'], vnaccs_row['GIO_DK_SUA'])
                safe_set_cell(ws1, 'R8', R8_value)
                safe_set_cell(ws1, 'R83', R8_value)
                safe_set_cell(ws2, 'R8', R8_value)
            elif 'NGAY_DK_SUA' in vnaccs_row:
                safe_set_cell(ws1, 'R8', format_date(vnaccs_row['NGAY_DK_SUA']))
                safe_set_cell(ws1, 'R83', format_date(vnaccs_row['NGAY_DK_SUA']))
                safe_set_cell(ws2, 'R8', format_date(vnaccs_row['NGAY_DK_SUA']))
            else:  # Nếu rỗng
                safe_set_cell(ws1, 'R8', '')
                safe_set_cell(ws1, 'R83', '')
                safe_set_cell(ws2, 'R8', '')
            
        # Từ bảng DTOKHAIMD_VNACCS2
        if vnaccs2_row is not None:
            if 'CANHAN_TOCHUC' in vnaccs2_row:
                safe_set_cell(ws1, 'P6', vnaccs2_row['CANHAN_TOCHUC'])
            
            if 'MA_HIEU_PTVC' in vnaccs2_row:
                safe_set_cell(ws1, 'P6', vnaccs2_row['MA_HIEU_PTVC'])
            
            if 'NHOM_HO_SO' in vnaccs2_row:
                safe_set_cell(ws1, 'AE7', vnaccs2_row['NHOM_HO_SO'])
                safe_set_cell(ws1, 'AE82', vnaccs2_row['NHOM_HO_SO'])
                safe_set_cell(ws2, 'AE7', vnaccs2_row['NHOM_HO_SO'])
            
            if 'DIA_CHI_DT1' in vnaccs2_row:
                safe_set_cell(ws1, 'H25', vnaccs2_row['DIA_CHI_DT1'])
            
            if 'DIA_CHI_DT2' in vnaccs2_row:
                safe_set_cell(ws1, 'U25', vnaccs2_row['DIA_CHI_DT2'])
            
            if 'DIA_CHI_DT3' in vnaccs2_row:
                safe_set_cell(ws1, 'H26', vnaccs2_row['DIA_CHI_DT3'])
            
            if 'DIA_CHI_DT4' in vnaccs2_row:
                safe_set_cell(ws1, 'U26', vnaccs2_row['DIA_CHI_DT4'])
            
            if 'MA_DD_LUU_KHO' in vnaccs2_row:
                safe_set_cell(ws1, 'U30', vnaccs2_row['MA_DD_LUU_KHO'])
            
            if 'CHI_THI_HQ_NGAY' in vnaccs2_row:
                chi_thi_ngay = vnaccs2_row.get('CHI_THI_HQ_NGAY')
                safe_set_cell(ws1, 'D90', format_date(chi_thi_ngay) if chi_thi_ngay else '')
                chi_thi_hq_ten = vnaccs2_row.get('CHI_THI_HQ_TEN') or ''
                safe_set_cell(ws1, 'I90', str(chi_thi_hq_ten).strip())
                chi_thi_hq_nd = vnaccs2_row.get('CHI_THI_HQ_ND') or ''
                safe_set_cell(ws1, 'R90', str(chi_thi_hq_nd).strip())
        
        # Ghi giá trị vào ô P6 theo format: A12  2 [ 4 ]
        # Trong đó MA_LH = A12, MA_HIEU_PTVC = 2, CANHAN_TOCHUC = 4
        p6_value = ""
        
        # Lấy MA_LH từ dtokhaimd_row
        if dtokhaimd_row is not None and 'MA_LH' in dtokhaimd_row:
            p6_value += str(dtokhaimd_row['MA_LH'])
        
        # Lấy MA_HIEU_PTVC từ vnaccs2_row
        if vnaccs2_row is not None and 'MA_HIEU_PTVC' in vnaccs2_row:
            p6_value += f"  {vnaccs2_row['MA_HIEU_PTVC']}"
        
        # Lấy CANHAN_TOCHUC từ vnaccs2_row
        if vnaccs2_row is not None and 'CANHAN_TOCHUC' in vnaccs2_row:
            p6_value += f" [ {vnaccs2_row['CANHAN_TOCHUC']} ]"
        
        # Ghi vào ô P6
        if p6_value:
            safe_set_cell(ws1, 'P6', p6_value)
            safe_set_cell(ws1, 'P81', p6_value)
            safe_set_cell(ws2, 'P6', p6_value)
        
        # Ghi giá trị vào ô J41 theo format: A - 250822-89A(CY)
        # Trong đó MA_HDTM = A, SO_HDTM = 250822-89A(CY)
        j41_value = ""
        
        if dtokhaimd_row is not None and 'MA_HDTM' in dtokhaimd_row:
            j41_value += str(dtokhaimd_row['MA_HDTM'])
        
        if dtokhaimd_row is not None and 'SO_HDTM' in dtokhaimd_row:
            j41_value += f" - {dtokhaimd_row['SO_HDTM']}"
        
        if j41_value:
            safe_set_cell(ws1, 'J41', j41_value)
        
        # Ghi giá trị vào ô J45 theo format: A - CFR - USD - 
        # Trong đó MA_GH = CFR, MA_NT = USD, MA_PL_GIA_HDTM = A
        j45_value = ""
        
        if dtokhaimd_row is not None and 'MA_PL_GIA_HDTM' in dtokhaimd_row:
            j45_value += str(dtokhaimd_row['MA_PL_GIA_HDTM'])
        
        if dtokhaimd_row is not None and 'MA_GH' in dtokhaimd_row:
            j45_value += f" - {dtokhaimd_row['MA_GH']}"
        
        if dtokhaimd_row is not None and 'MA_NT' in dtokhaimd_row:
            j45_value += f" - {dtokhaimd_row['MA_NT']} -"
        
        if j45_value:
            safe_set_cell(ws1, 'J45', j45_value)
        
        # Ghi giá trị vào ô K84 theo format: ETC - 722425995230
        # Trong đó MA_LOAI_DINH_KEM = ETC, SO_DINH_KEM = 722425995230
        k84_value = ""

        if vnaccs_row is not None and 'MA_LOAI_DINH_KEM' in vnaccs_row:
            k84_value += str(vnaccs_row['MA_LOAI_DINH_KEM'])

        if vnaccs_row is not None and 'SO_DINH_KEM' in vnaccs_row and vnaccs_row['SO_DINH_KEM'] is not None:
            k84_value += f" - {str(int(vnaccs_row['SO_DINH_KEM']))}"
        else:
            k84_value = " - "

        if k84_value:
            safe_set_cell(ws1, 'K84', k84_value)
        
        s84_value = ""
        if vnaccs_row is not None and 'MA_LOAI_DINH_KEM2' in vnaccs_row:
            s84_value += str(vnaccs_row['MA_LOAI_DINH_KEM2'])
        if vnaccs_row is not None and 'SO_DINH_KEM2' in vnaccs_row and vnaccs_row['SO_DINH_KEM2'] is not None:
            s84_value += f" - {str(int(vnaccs_row['SO_DINH_KEM2']))}"
        else:
            s84_value = " - "
        if s84_value:
            safe_set_cell(ws1, 'S84', s84_value)
        
        z84_value = ""  
        if vnaccs_row is not None and 'MA_LOAI_DINH_KEM3' in vnaccs_row:
            z84_value += str(vnaccs_row['MA_LOAI_DINH_KEM3'])
        if vnaccs_row is not None and 'SO_DINH_KEM3' in vnaccs_row and vnaccs_row['SO_DINH_KEM3'] is not None:
            z84_value += f" - {str(int(vnaccs_row['SO_DINH_KEM3']))}"
        else:
            z84_value = " - "
        if z84_value:
            safe_set_cell(ws1, 'Z84', z84_value)
        
        # Copy ws2 vào ws1
        #dongcuoi = 139
        # Tắt FitToPage để Excel không tự tính lại ngắt trang; dùng scale cố định để vừa trang
        ws1.sheet_properties.pageSetUpPr.fitToPage = True
        ws1.page_setup.fitToWidth = 1
        ws1.page_setup.fitToHeight = 0
        ws1.page_setup.scale = None
        
        # Xóa tiêu đề lặp lại theo trang (nếu có)
        try:
            ws1.print_title_rows = None
            ws1.print_title_cols = None
        except Exception:
            pass
        
        # Xóa toàn bộ ngắt trang cũ trước khi thiết lập lại
        try:
            from openpyxl.worksheet.pagebreak import PageBreak
            ws1.row_breaks = PageBreak()
            ws1.col_breaks = PageBreak()
        except Exception:
            pass
        
        # Chuẩn bị và sao chép các trang chi tiết hàng
        goods_df = data.get('dhangmddk', pd.DataFrame())
        if not isinstance(goods_df, pd.DataFrame):
            goods_df = pd.DataFrame(goods_df or [])
        if not goods_df.empty:
            goods_df = goods_df.reset_index(drop=True)
        goods_row_count = len(goods_df)

        declared_goods_pages = 0
        if dtokhaimd_row is not None and 'SOHANG' in dtokhaimd_row:
            try:
                declared_goods_pages = int(dtokhaimd_row['SOHANG'])
            except (TypeError, ValueError):
                declared_goods_pages = 0

        so_trang_phu = max(goods_row_count, declared_goods_pages)
        total_pages = total_pages_declared if total_pages_declared and total_pages_declared >= 2 else max(2, 2 + so_trang_phu)

        safe_set_cell(ws1, 'AF1', f"1/{total_pages}")
        safe_set_cell(ws1, 'AF76', f"2/{total_pages}")

        if goods_row_count:
            for page_idx in range(goods_row_count):
                page_no = min(total_pages, 3 + page_idx)
                clear_cells_by_mapping(ws2, TKN_DETAIL_CLEAR_CELLS, safe_set_cell)
                safe_set_cell(ws2, 'AE30', '')
                goods_row = goods_df.iloc[page_idx]
                populate_worksheet_from_mapping(ws2, goods_row, TKN_DETAIL_FIELD_MAP, safe_set_cell)
                thuekhac_luong_tt = goods_row.get('THUEKHAC_LUONG_TT') if hasattr(goods_row, 'get') else goods_row['THUEKHAC_LUONG_TT'] if 'THUEKHAC_LUONG_TT' in goods_row else None
                has_thuekhac_luong_tt = False
                if thuekhac_luong_tt is not None:
                    if isinstance(thuekhac_luong_tt, str):
                        has_thuekhac_luong_tt = thuekhac_luong_tt.strip() != ''
                    else:
                        try:
                            has_thuekhac_luong_tt = not pd.isna(thuekhac_luong_tt)
                        except Exception:
                            has_thuekhac_luong_tt = True
                if has_thuekhac_luong_tt:
                    thuekhac_ma_dvt = goods_row.get('THUEKHAC_MA_DVT_CHUAN') if hasattr(goods_row, 'get') else goods_row['THUEKHAC_MA_DVT_CHUAN'] if 'THUEKHAC_MA_DVT_CHUAN' in goods_row else None
                    if thuekhac_ma_dvt is not None:
                        safe_set_cell(ws2, 'AE30', str(thuekhac_ma_dvt).strip())
                safe_set_cell(ws2, 'AF1', f"{page_no}/{total_pages}")
                start_row = TKN_DETAIL_FIRST_DETAIL_ROW + page_idx * TKN_DETAIL_PAGE_HEIGHT
                copy_row_range_with_style(ws2, ws1, TKN_DETAIL_TEMPLATE_RANGE, start_row)
        else:
            clear_cells_by_mapping(ws2, TKN_DETAIL_CLEAR_CELLS, safe_set_cell)
            safe_set_cell(ws2, 'AE30', '')

        if so_trang_phu > goods_row_count:
            for extra_idx in range(goods_row_count, so_trang_phu):
                page_no = min(total_pages, 3 + extra_idx)
                clear_cells_by_mapping(ws2, TKN_DETAIL_CLEAR_CELLS, safe_set_cell)
                safe_set_cell(ws2, 'AE30', '')
                safe_set_cell(ws2, 'AF1', f"{page_no}/{total_pages}")
                start_row = TKN_DETAIL_FIRST_DETAIL_ROW + extra_idx * TKN_DETAIL_PAGE_HEIGHT
                copy_row_range_with_style(ws2, ws1, TKN_DETAIL_TEMPLATE_RANGE, start_row)
        # Thiết lập khu vực in theo tổng số hàng (header 138 hàng + n*53)
        tong_hang = 138 + so_trang_phu * TKN_DETAIL_PAGE_HEIGHT
        ws1.print_area = f"A1:AH{tong_hang}"
        ws1.print_options.horizontalCentered = True
        
        # Thiết lập ngắt trang: trang 1 tại 75, trang 2 tại 138; các trang sau: 138 + i*53
        try:
            break_positions = []
            ws1.row_breaks.append(Break(id=75))
            break_positions.append(75)
            ws1.row_breaks.append(Break(id=138))
            break_positions.append(138)
            for i in range(1, so_trang_phu + 1):
                r = 138 + i * TKN_DETAIL_PAGE_HEIGHT
                ws1.row_breaks.append(Break(id=r))
                break_positions.append(r)
            if break_positions:
                print(f"Manual page breaks at rows: {break_positions}")
        except Exception:
            pass
       
        if ws2.title in wb.sheetnames:
            wb.remove(ws2)
        try:
            ws1.title = 'TKN'
        except Exception:
            ws1.title = ws1.title  # giữ nguyên nếu rename thất bại

        wb.save(output_file)
        wb.close()
        print(f"✓ Đã tạo file TKN: {output_file}")
        
    except Exception as e:
        print(f"✗ Lỗi khi điền dữ liệu vào Excel: {e}")
        messagebox.showerror('Excel Error', f'Lỗi khi điền dữ liệu: {e}') 
        

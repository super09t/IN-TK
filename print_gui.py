# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from datetime import datetime
import threading
import os
import time
import subprocess
import shutil
import csv
import sys
import importlib.util
import json
try:
    import ocr_easy as _ocr_easy_module
except Exception:
    _ocr_easy_module = None
try:
    import ocr_paddle as _ocr_paddle_module
except Exception:
    _ocr_paddle_module = None
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

# Import các hàm backend
from sql_helpers_new import (
    SQL_Select_query_params,
    get_cd_details_bulk,
    PrintCD_TKN,
    PrintCD_TKX,
    normalize_dtokhaimdid,
)
from sql_helpers_new import select_query_df_pyodbc

COLUMNS = [
    'id', 'imex', 'so_tk', 'ma_lh', 'ma_hq', 'ngay_dk', 'invoice', 'invoice_date', 'status',
    'so_tk_dau_tien', 'van_don', 'so_hd', 'so_gp', 'dv_dt',
]
HEADERS = {
    'sel': 'Sel',
    'id': 'ID',
    'imex': 'IM/EX',
    'so_tk': 'SO_TK',
    'ma_lh': 'MA_LH',
    'ma_hq': 'MA_HQ',
    'ngay_dk': 'NGAY_DK',
    'invoice': 'Invoice',
    'invoice_date': 'Invoice_date',
    'status': 'Status',
    'so_tk_dau_tien': 'SOTK_FIRST',
    'van_don': 'VAN_DON',
    'so_hd': 'SO_HD',
    'so_gp': 'SO_GP',
    'dv_dt': 'DV_DT',
}
HIDDEN_COLUMNS = {
    'imex', 'invoice', 'invoice_date', 'status',
    'so_tk_dau_tien', 'van_don', 'so_hd', 'so_gp', 'dv_dt',
}
SEARCH_FIELDS = [
    ('id', 'ID'),
    ('so_tk', 'So to khai'),
    ('ma_hq', 'Ma hai quan'),
    ('so_tk_dau_tien', 'So to khai dau tien'),
    ('van_don', 'Van don'),
    ('so_hd', 'Hoa don'),
    ('so_gp', 'Giay phep'),
    ('dv_dt', 'Doi tac'),
]
SEARCH_FIELD_TO_DB = {
    'id': '_DTOKHAIMDID',
    'so_tk': 'SOTK',
    'ma_hq': 'MA_HQ',
    'so_tk_dau_tien': 'SOTK_DAU_TIEN',
    'van_don': 'VAN_DON',
    'so_hd': 'SO_HD',
    'so_gp': 'SO_GP',
    'dv_dt': 'DV_DT',
}
DEFAULT_DSHH_PATH = os.path.join(os.path.abspath('Excel_Tempt'), 'DSHH.xlsx')
DEFAULT_COLUMN_WIDTH = 110
COLUMN_WIDTHS = {
    'sel': 50,
    'id': 47,
    'imex': 70,
    'so_tk': 140,
    'ma_lh': 90,
    'ma_hq': 90,
    'ngay_dk': 115,
    'invoice': 120,
    'invoice_date': 130,
    'status': 110,
    'so_tk_dau_tien': 130,
    'van_don': 140,
    'so_hd': 130,
    'so_gp': 130,
    'dv_dt': 140,
}
TREEVIEW_EXTRA_PADDING = 30
CHECK_MARK = '\u2713'
STATUS_MAP = {
    'N': 'TK nhap',
    'E': 'TK truyen nhap',
    'P': 'TK phan luong',
    'T': 'TKTQ',
}

def _detect_paddle_available():
    try:
        return importlib.util.find_spec('paddleocr') is not None
    except Exception:
        return False


def _detect_easy_available():
    try:
        return importlib.util.find_spec('easyocr') is not None
    except Exception:
        return False


class PrintApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('AUTO GUI')
        self.geometry('1000x700')
        base_dir = os.path.dirname(os.path.abspath(__file__))
        self.config_path = os.path.join(base_dir, 'config.json')
        self.app_config = self._load_config()
        self.output_dir = os.path.abspath(self.app_config.get('output_dir', 'output'))
        os.makedirs(self.output_dir, exist_ok=True)
        self.default_dshh_path = DEFAULT_DSHH_PATH
        self.ent_dshh = None
        self.use_dshh_var = tk.BooleanVar(value=False)
        self.kv1_full_data = []
        self.search_entries = {}
        self.ids_tree = None
        self.ids_data = []
        self.ids_fetch_thread = None

        self._build_tabs()

    def _load_config(self):
        try:
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if isinstance(data, dict):
                        return data
        except Exception as exc:
            print(f'Warning: không thể đọc config: {exc}')
        return {}

    def _build_tabs(self):
        """Tạo 3 tab chính"""
        # Tạo notebook (tab container)
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill='both', expand=True, padx=8, pady=8)
        
        # Tab 1: CSDL
        self.tab_csdl = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_csdl, text='CSDL')
        self._build_tab_csdl()
        
        # Tab 2: In tờ khai
        self.tab_print = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_print, text='In tờ khai')
        self._build_tab_print()
        
        # Tab 3: Số định danh
        self.tab_ids = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_ids, text='Số định danh')
        self._build_tab_ids()

    def _build_tab_csdl(self):
        """Tab CSDL - cấu hình kết nối database"""
        # Frame chính
        main_frame = ttk.Frame(self.tab_csdl)
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Tiêu đề
        title_label = ttk.Label(main_frame, text='Cấu hình kết nối cơ sở dữ liệu', 
                               font=('Segoe UI', 14, 'bold'))
        title_label.pack(pady=(0, 20))
        
        # Frame chứa các trường nhập liệu
        input_frame = ttk.LabelFrame(main_frame, text='Thông tin kết nối', padding=20)
        input_frame.pack(fill='x', pady=(0, 20))
        
        # Server
        ttk.Label(input_frame, text='Server:', font=('Segoe UI', 10)).grid(row=0, column=0, sticky='w', pady=5)
        self.ent_server = ttk.Entry(input_frame, width=30, font=('Segoe UI', 10))
        self.ent_server.grid(row=0, column=1, padx=(10, 0), pady=5, sticky='ew')
        self.ent_server.insert(0, self.app_config.get('server', '192.168.100.6,1433'))
        
        # Database
        ttk.Label(input_frame, text='Database:', font=('Segoe UI', 10)).grid(row=1, column=0, sticky='w', pady=5)
        self.ent_db = ttk.Entry(input_frame, width=30, font=('Segoe UI', 10))
        self.ent_db.grid(row=1, column=1, padx=(10, 0), pady=5, sticky='ew')
        self.ent_db.insert(0, self.app_config.get('database', 'Ecus5vnaccs_liem'))
        
        # Username
        ttk.Label(input_frame, text='Username:', font=('Segoe UI', 10)).grid(row=2, column=0, sticky='w', pady=5)
        self.ent_user = ttk.Entry(input_frame, width=30, font=('Segoe UI', 10))
        self.ent_user.grid(row=2, column=1, padx=(10, 0), pady=5, sticky='ew')
        self.ent_user.insert(0, self.app_config.get('username', 'sa1'))
        
        # Password
        ttk.Label(input_frame, text='Password:', font=('Segoe UI', 10)).grid(row=3, column=0, sticky='w', pady=5)
        self.ent_pwd = ttk.Entry(input_frame, width=30, font=('Segoe UI', 10), show='*')
        self.ent_pwd.grid(row=3, column=1, padx=(10, 0), pady=5, sticky='ew')
        self.ent_pwd.insert(0, self.app_config.get('password', '12345678sa'))
        
        # MA_DV
        ttk.Label(input_frame, text='MA_DV:', font=('Segoe UI', 10)).grid(row=4, column=0, sticky='w', pady=5)
        self.ent_madv = ttk.Entry(input_frame, width=30, font=('Segoe UI', 10))
        self.ent_madv.grid(row=4, column=1, padx=(10, 0), pady=5, sticky='ew')
        self.ent_madv.insert(0, self.app_config.get('madv', '0314404243001'))
        
        # Cấu hình grid weights
        input_frame.grid_columnconfigure(1, weight=1)
        
        # Nút test kết nối
        test_frame = ttk.Frame(main_frame)
        test_frame.pack(fill='x', pady=(0, 20))
        
        ttk.Button(test_frame, text='Test kết nối', command=self._test_connection, 
                  style='Accent.TButton').pack(side='left')
        ttk.Button(test_frame, text='Lưu cấu hình', command=self.on_save_config).pack(side='left', padx=(10, 0))
        
        # Thông báo kết quả test
        self.lbl_connection_status = ttk.Label(test_frame, text='', font=('Segoe UI', 10))
        self.lbl_connection_status.pack(side='left', padx=(10, 0))

        node_frame = ttk.LabelFrame(main_frame, text='Cấu hình Node.js', padding=12)
        node_frame.pack(fill='x', pady=(0, 20))
        ttk.Label(node_frame, text='Node.exe').pack(side='left')
        self.ent_node = ttk.Entry(node_frame, width=50)
        common_node = os.path.join('C:\\Program Files', 'nodejs', 'node.exe')
        detected_node = self.app_config.get('node_path') or shutil.which('node') or shutil.which('node.exe') or (common_node if os.path.exists(common_node) else '')
        self.ent_node.insert(0, detected_node)
        self.ent_node.pack(side='left', padx=6, fill='x', expand=True)
        ttk.Button(node_frame, text='Chọn Node.exe', command=self.on_choose_node).pack(side='left', padx=(6, 0))
        self.lbl_save_status = ttk.Label(main_frame, text='', foreground='#0066cc', font=('Segoe UI', 9))
        self.lbl_save_status.pack(anchor='w', pady=(0, 10))

    def _build_tab_print(self):
        """Tab In tờ khai - chứa các chức năng in tờ khai"""
        # Frame chính
        main_frame = ttk.Frame(self.tab_print)
        main_frame.pack(fill='both', expand=True, padx=8, pady=8)
        
        # Gọi các hàm build cũ nhưng với parent là main_frame
        self._build_controls(main_frame)
        self._build_tables(main_frame)

    def _build_tab_ids(self):
        """Tab lấy số định danh"""
        main_frame = ttk.Frame(self.tab_ids)
        main_frame.pack(fill='both', expand=True, padx=16, pady=16)

        header = ttk.Label(main_frame, text='Lấy số định danh', font=('Segoe UI', 14, 'bold'))
        header.pack(anchor='w', pady=(0, 12))

        form = ttk.Frame(main_frame)
        form.pack(fill='x', pady=(0, 12))

        ttk.Label(form, text='Số lượng:').grid(row=0, column=0, sticky='w')
        self.ent_ids_count = ttk.Entry(form, width=8)
        self.ent_ids_count.insert(0, '1')
        self.ent_ids_count.grid(row=0, column=1, padx=(6, 12))

        ttk.Label(form, text='Mã DN (user):').grid(row=0, column=2, sticky='w')
        self.ent_ids_user = ttk.Entry(form, width=20)
        default_user = getattr(self, 'ent_madv', None)
        if default_user is not None:
            self.ent_ids_user.insert(0, default_user.get().strip())
        self.ent_ids_user.grid(row=0, column=3, padx=(6, 12))

        ttk.Button(form, text='Lấy số định danh', command=self.on_fetch_ids).grid(row=0, column=4, padx=(0, 12))

        self.lbl_ids_status = ttk.Label(main_frame, text='', foreground='#0066cc')
        self.lbl_ids_status.pack(anchor='w', pady=(0, 8))

        columns = ('idx', 'code', 'timestamp')
        tree = ttk.Treeview(main_frame, columns=columns, show='headings', height=18, selectmode='extended')
        tree.heading('idx', text='STT')
        tree.heading('code', text='Số định danh')
        tree.heading('timestamp', text='Thời gian')
        tree.column('idx', width=80, anchor='center')
        tree.column('code', width=220, anchor='center')
        tree.column('timestamp', width=180, anchor='center')
        tree.pack(fill='both', expand=True)
        tree.bind('<Control-c>', lambda e: self.on_copy_ids(from_event=True))
        tree.bind('<Control-C>', lambda e: self.on_copy_ids(from_event=True))
        self.ids_tree = tree

    def _test_connection(self):
        """Test kết nối database"""
        def test_task():
            try:
                Sqlhost = self._get_sqlhost()
                # Test query đơn giản
                df = select_query_df_pyodbc(Sqlhost, "SELECT 1 as test")
                if df is not None and not df.empty:
                    self.after(0, lambda: self.lbl_connection_status.config(
                        text='✅ Kết nối thành công!', foreground='green'))
                else:
                    self.after(0, lambda: self.lbl_connection_status.config(
                        text='❌ Kết nối thất bại!', foreground='red'))
            except Exception as e:
                self.after(0, lambda: self.lbl_connection_status.config(
                    text=f'❌ Lỗi: {str(e)[:50]}...', foreground='red'))
        
        self.lbl_connection_status.config(text='🔄 Đang test...', foreground='blue')
        threading.Thread(target=test_task, daemon=True).start()

    def on_save_config(self):
        data = {
            'server': (self.ent_server.get() or '').strip(),
            'database': (self.ent_db.get() or '').strip(),
            'username': (self.ent_user.get() or '').strip(),
            'password': self.ent_pwd.get(),
            'madv': (self.ent_madv.get() or '').strip(),
            'node_path': (self.ent_node.get() or '').strip(),
            'output_dir': self.output_dir,
        }
        try:
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            self.app_config = data
            self.lbl_save_status.config(text='Đã lưu cấu hình.', foreground='#006600')
        except Exception as exc:
            self.lbl_save_status.config(text='Không lưu được cấu hình.', foreground='#cc0000')
            messagebox.showerror('Lỗi', f'Không thể lưu cấu hình: {exc}')

    def _build_controls(self, parent=None):
        if parent is None:
            parent = self
            
        frm = ttk.Frame(parent)
        frm.pack(fill='x', padx=8, pady=8)

        # Hàng lọc
        ttk.Label(frm, text='Từ ngày (yyyy-mm-dd)').grid(row=0, column=0, sticky='w', pady=(6,0))
        self.ent_from = ttk.Entry(frm, width=18)
        self.ent_from.grid(row=0, column=1, padx=6, pady=(6,0))

        ttk.Label(frm, text='Đến ngày').grid(row=0, column=2, sticky='w', pady=(6,0))
        self.ent_to = ttk.Entry(frm, width=18)
        self.ent_to.grid(row=0, column=3, padx=6, pady=(6,0))

        ttk.Label(frm, text='Loại hình').grid(row=0, column=4, sticky='w', pady=(6,0))
        self.cbo_lh = ttk.Combobox(frm, width=14, state='readonly', values=['', 'B11', 'B13', 'A11', 'A12', 'A41'])
        self.cbo_lh.current(0)
        self.cbo_lh.grid(row=0, column=5, padx=6, pady=(6,0))

        ttk.Label(frm, text='IM/EX').grid(row=0, column=6, sticky='w', pady=(6,0))
        self.cbo_imex = ttk.Combobox(frm, width=10, state='readonly', values=['', 'N', 'X'])
        self.cbo_imex.current(0)
        self.cbo_imex.grid(row=0, column=7, padx=6, pady=(6,0))

        # Nút chức năng
        btns = ttk.Frame(parent)
        btns.pack(fill='x', padx=8, pady=4)

        ttk.Button(btns, text='Getdata', command=self.on_getdata).pack(side='left', padx=4)
        ttk.Button(btns, text='Data_down', command=self.on_data_down).pack(side='left', padx=4)
        ttk.Button(btns, text='Data_up', command=self.on_data_up).pack(side='left', padx=4)
        ttk.Button(btns, text='Chọn Folder', command=self.on_choose_folder).pack(side='left', padx=4)
        ttk.Button(btns, text='Xuất tờ khai excel', command=self.on_export).pack(side='left', padx=4)
        ttk.Button(btns, text='Xuất Mã Vạch', command=self.on_export_mv).pack(side='left', padx=4)

        self.lbl_out = ttk.Label(btns, text=f'Lưu tại: {self.output_dir}', foreground='#0066cc', cursor='hand2')
        self.lbl_out.pack(side='left', padx=10)
        self.lbl_out.bind('<Button-1>', lambda e: self._open_output_dir())

        dshh_frame = ttk.Frame(parent)
        dshh_frame.pack(fill='x', padx=8, pady=(0,8))
        ttk.Label(dshh_frame, text='DSHH.xlsx').pack(side='left')
        self.ent_dshh = ttk.Entry(dshh_frame, width=25)
        default_dshh = self.default_dshh_path or ''
        if default_dshh:
            self.ent_dshh.insert(0, default_dshh)
        self.ent_dshh.pack(side='left', padx=6)
        ttk.Button(dshh_frame, text='Chọn file DSHH', command=self.on_choose_dshh).pack(side='left', padx=(0, 6))
        ttk.Checkbutton(
            dshh_frame,
            text='Dùng file này',
            variable=self.use_dshh_var
        ).pack(side='left', padx=(0, 6))
        lbl_template = ttk.Label(dshh_frame, text='Tải template DSHH', foreground='#0066cc', cursor='hand2')
        lbl_template.pack(side='left')
        lbl_template.bind('<Button-1>', lambda e: self.on_open_dshh_template())

    def _build_tables(self, parent=None):
        if parent is None:
            parent = self
            
        # Khu vực kv1/kv2
        container = ttk.Frame(parent)
        container.pack(fill='both', expand=True, padx=8, pady=8)
        container.grid_rowconfigure(0, weight=1)
        container.grid_rowconfigure(1, weight=1)
        container.grid_columnconfigure(0, weight=0)
        container.grid_columnconfigure(1, weight=1)

        top = ttk.Frame(container, relief='groove', padding=6)
        top.grid(row=0, column=0, sticky='nsw', pady=(0,4))
        self.kv1 = self._create_table(top, 'KV1 - Dữ liệu nguồn')

        bottom = ttk.Frame(container, relief='groove', padding=6)
        bottom.grid(row=1, column=0, sticky='nsw', pady=(4,0))
        self.kv2 = self._create_table(bottom, 'KV2 - Danh sách sẽ xuất')

        self._build_search_panel(container)

    def _create_table(self, parent, title):
        frame = ttk.Frame(parent)
        
        # Header với title, checkbox tích toàn bộ, số lượng và trạng thái
        header_frame = ttk.Frame(frame)
        header_frame.pack(fill='x', pady=(0,4))
        
        ttk.Label(header_frame, text=title, font=('Segoe UI', 10, 'bold')).pack(side='left')
        
        # Checkbox tích toàn bộ
        select_all_var = tk.BooleanVar()
        select_all_cb = ttk.Checkbutton(header_frame, text='Tích toàn bộ', variable=select_all_var)
        select_all_cb.pack(side='left', padx=(10, 0))
        
        # Bind command sau khi tạo treeview
        def create_select_all_command(tree, var):
            def command():
                self._on_select_all_toggle(tree, var)
            return command
        
        # Label hiển thị tổng số dòng
        count_label = ttk.Label(header_frame, text='(0 dòng)', foreground='#666', font=('Segoe UI', 9))
        count_label.pack(side='left', padx=(10, 0))
        
        # Label hiển thị trạng thái (chỉ cho KV2)
        status_label = ttk.Label(header_frame, text='', foreground='#0066cc', font=('Segoe UI', 9, 'bold'))
        status_label.pack(side='left', padx=(10, 0))

        cols = ['sel'] + COLUMNS
        visible_cols = [c for c in cols if c not in HIDDEN_COLUMNS]
        tv = ttk.Treeview(frame, columns=cols, show='headings', selectmode='none', height=22)
        tv.configure(displaycolumns=visible_cols)

        total_width = 0
        for c in cols:
            tv.heading(c, text=HEADERS.get(c, c))
            width = COLUMN_WIDTHS.get(c, DEFAULT_COLUMN_WIDTH)
            anchor = 'center' if c == 'sel' else 'w'
            tv.column(c, width=width, anchor=anchor, stretch=False)
            if c in visible_cols:
                total_width += width
        total_width += TREEVIEW_EXTRA_PADDING

        # Scrollbars
        yscroll = ttk.Scrollbar(frame, orient='vertical', command=tv.yview)
        xscroll = ttk.Scrollbar(frame, orient='horizontal', command=tv.xview)
        tv.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)

        # Pack the container frame so it becomes visible
        frame.pack(side='left', fill='y', expand=False)
        frame.pack_propagate(False)
        frame.config(width=total_width)
        tv.pack(fill='both', expand=True)
        xscroll.pack(fill='x')
        yscroll.place(relx=1.0, rely=0.0, relheight=1.0, anchor='ne')

        # Toggle "checkbox" khi click vào cột Sel
        tv.bind('<Button-1>', lambda e, tree=tv: self._on_click_tree(e, tree))
        
        # Lưu reference đến count_label, status_label và select_all_var để có thể cập nhật sau
        tv.count_label = count_label
        tv.status_label = status_label
        tv.select_all_var = select_all_var
        
        # Bind command cho checkbox "Tích toàn bộ" sau khi tạo treeview
        select_all_cb.config(command=create_select_all_command(tv, select_all_var))
        
        # Khởi tạo số lượng ban đầu
        self._update_count(tv)
        
        return tv

    def _build_search_panel(self, container):
        panel = ttk.LabelFrame(container, text='Tim kiem', padding=10)
        panel.grid(row=0, column=1, rowspan=2, sticky='ns', padx=(10, 0))
        panel.columnconfigure(1, weight=0)
        self.search_entries = {}
        for idx, (field, label) in enumerate(SEARCH_FIELDS):
            ttk.Label(panel, text=label).grid(row=idx, column=0, sticky='w', pady=3, padx=(0, 6))
            entry = ttk.Entry(panel, width=16)
            entry.grid(row=idx, column=1, sticky='w', pady=3)
            self.search_entries[field] = entry
        ttk.Button(panel, text='Xoa tim kiem', command=self._reset_search_filters).grid(
            row=len(SEARCH_FIELDS), column=0, columnspan=2, sticky='ew', pady=(10, 0)
        )

    def _update_count(self, tree):
        """Cập nhật số lượng dòng trong tree"""
        count = len(tree.get_children())
        tree.count_label.config(text=f'({count} dòng)')

    def _update_status(self, tree, status):
        """Cập nhật trạng thái cho tree (chỉ KV2)"""
        if hasattr(tree, 'status_label'):
            if status == 'loading':
                tree.status_label.config(text='Đang lấy mã vạch...', foreground='#ff6600')
            elif status == 'completed':
                tree.status_label.config(text='Hoàn thành', foreground='#006600')
            elif status == 'error':
                tree.status_label.config(text='Lỗi', foreground='#cc0000')
            elif status == 'clear':
                tree.status_label.config(text='', foreground='#0066cc')

    def _on_click_tree(self, event, tree):
        region = tree.identify('region', event.x, event.y)
        if region != 'cell':
            return
        row_id = tree.identify_row(event.y)
        col = tree.identify_column(event.x)
        if not row_id or col != '#1':  # '#1' là cột sel
            return
        cur = tree.set(row_id, 'sel')
        tree.set(row_id, 'sel', CHECK_MARK if cur != CHECK_MARK else '')
        # Cập nhật trạng thái checkbox "Tích toàn bộ"
        self._update_select_all_checkbox(tree)

    def _get_sqlhost(self):
        host = self.ent_server.get().strip()
        db = self.ent_db.get().strip()
        user = self.ent_user.get().strip()
        pwd = self.ent_pwd.get().strip()
        if not all([host, db, user, pwd]):
            raise ValueError('Vui lòng nhập đủ Server/Database/User/Password')
        return (host, db, user, pwd)

    def on_getdata(self):
        def task():
            try:
                Sqlhost = self._get_sqlhost()
                ma_dv = self.ent_madv.get().strip()
                base_sql = (
                    "SELECT TOP 20 _DTOKHAIMDID AS ID, _XORn AS [IM/EX], SOTK AS SO_TK, MA_LH, MA_HQ, NGAY_DK, "
                    "SO_HDTM AS Invoice, NGAY_HDTM AS Invoice_date, TTTK AS Status, "
                    "SOTK_DAU_TIEN, VAN_DON, SO_HD, SO_GP, DV_DT "
                    "FROM DTOKHAIMD WHERE MA_DV = ?"
                )
                params = [ma_dv]
                from_str = self.ent_from.get().strip()
                to_str = self.ent_to.get().strip()
                lh = self.cbo_lh.get().strip()
                imex = self.cbo_imex.get().strip().upper()
                status = 'T'
                # Bổ sung điều kiện theo ngày
                if from_str and to_str:
                    base_sql += " AND NGAY_DK BETWEEN ? AND ?"
                    params.extend([from_str, to_str])
                elif from_str:
                    base_sql += " AND NGAY_DK >= ?"
                    params.append(from_str)
                elif to_str:
                    base_sql += " AND NGAY_DK <= ?"
                    params.append(to_str)
                # Bổ sung điều kiện loại hình
                if lh:
                    base_sql += " AND MA_LH = ?"
                    params.append(lh)
                # Bổ sung điều kiện IM/EX
                if imex:
                    base_sql += " AND _XORn = ?"
                    params.append(imex)
                # Bổ sung điều kiện Status
                if status:
                    base_sql += " AND TTTK = ?"
                    params.append(status)
                search_criteria = self._get_search_criteria()
                for key, text in search_criteria.items():
                    column = SEARCH_FIELD_TO_DB.get(key)
                    if column:
                        base_sql += f" AND {column} LIKE ?"
                        params.append(f"%{text}%")
                base_sql += " ORDER BY NGAY_DK DESC"
                df = select_query_df_pyodbc(Sqlhost, base_sql, params=tuple(params))
                self.after(0, lambda d=df: self._fill_table(self.kv1, d))
                self.after(0, lambda d=df: messagebox.showinfo('Thông báo', f'Hoàn thành! Số dòng: {0 if d is None else len(d)}'))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror('Lỗi', f'Getdata lỗi: {e}'))
        threading.Thread(target=task, daemon=True).start()

    def _fill_table(self, tree, df):
        rows = []
        if df is None or df.empty:
            self._render_rows(tree, rows)
            if tree is self.kv1:
                self.kv1_full_data = []
            return
        for _, row in df.iterrows():
            rows.append(self._build_row_data(row))
        if tree is self.kv1:
            self.kv1_full_data = [r.copy() for r in rows]
            if self._has_search_filters():
                self._apply_search_filters()
                return
        self._render_rows(tree, rows)

    def _build_row_data(self, row):
        raw_status = str(row.get('Status', '') or '')
        mapped_status = STATUS_MAP.get(raw_status.strip().upper(), raw_status)
        data = {
            'id': str(row.get('ID', '') or ''),
            'imex': str(row.get('IM/EX', '') or ''),
            'so_tk': str(row.get('SO_TK', '') or ''),
            'ma_lh': str(row.get('MA_LH', '') or ''),
            'ma_hq': str(row.get('MA_HQ', '') or ''),
            'ngay_dk': self._fmt_date(row.get('NGAY_DK')),
            'invoice': str(row.get('Invoice', '') or ''),
            'invoice_date': self._fmt_date(row.get('Invoice_date')),
            'status': mapped_status,
            'so_tk_dau_tien': str(row.get('SOTK_DAU_TIEN', '') or ''),
            'van_don': str(row.get('VAN_DON', '') or ''),
            'so_hd': str(row.get('SO_HD', '') or ''),
            'so_gp': str(row.get('SO_GP', '') or ''),
            'dv_dt': str(row.get('DV_DT', '') or ''),
        }
        return data

    def _render_rows(self, tree, rows):
        tree.delete(*tree.get_children())
        for row_data in rows:
            values = [''] + [row_data.get(col, '') for col in COLUMNS]
            tree.insert('', 'end', values=values)
        self._update_count(tree)
        self._update_select_all_checkbox(tree)

    def _fmt_date(self, v):
        if v is None or v == '':
            return ''
        try:
            if isinstance(v, str):
                return v.split(' ')[0]
            return v.strftime('%Y-%m-%d')
        except Exception:
            return str(v)

    def _format_mv_date(self, raw_value):
        if not raw_value:
            return ''
        if isinstance(raw_value, datetime):
            return raw_value.strftime('%d/%m/%Y')
        snippet = str(raw_value).strip()
        if not snippet:
            return ''
        snippet = snippet[:10]
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(snippet, fmt).strftime('%d/%m/%Y')
            except Exception:
                continue
        return snippet

    def _values_to_row_data(self, values):
        data = {}
        for idx, column in enumerate(COLUMNS, start=1):
            data[column] = values[idx] if idx < len(values) else ''
        return data

    def _add_row_to_cache(self, row_data):
        if not isinstance(self.kv1_full_data, list):
            self.kv1_full_data = []
        self.kv1_full_data.append(row_data.copy())

    def _rows_equal(self, left, right):
        for col in COLUMNS:
            if (left.get(col) or '') != (right.get(col) or ''):
                return False
        return True

    def _remove_row_from_cache(self, row_data):
        if not isinstance(self.kv1_full_data, list):
            self.kv1_full_data = []
        for idx, existing in enumerate(self.kv1_full_data):
            if self._rows_equal(existing, row_data):
                del self.kv1_full_data[idx]
                break

    def _get_search_criteria(self, lowercase=False):
        if not self.search_entries:
            return {}
        criteria = {}
        for key, entry in self.search_entries.items():
            text = entry.get().strip()
            if text:
                criteria[key] = text.lower() if lowercase else text
        return criteria

    def _has_search_filters(self):
        return bool(self._get_search_criteria())

    def _apply_search_filters(self, event=None):
        if not hasattr(self, 'kv1'):
            return
        rows = getattr(self, 'kv1_full_data', []) or []
        criteria = self._get_search_criteria(lowercase=True)
        if not criteria:
            target_rows = rows
        else:
            target_rows = []
            for row in rows:
                match = True
                for key, needle in criteria.items():
                    target = (row.get(key) or '').lower()
                    if needle not in target:
                        match = False
                        break
                if match:
                    target_rows.append(row)
        self._render_rows(self.kv1, target_rows)

    def _reset_search_filters(self):
        for entry in self.search_entries.values():
            entry.delete(0, tk.END)
        self._refresh_kv1_view()

    def _refresh_kv1_view(self):
        if not hasattr(self, 'kv1'):
            return
        if self._has_search_filters():
            self._apply_search_filters()
        else:
            rows = getattr(self, 'kv1_full_data', []) or []
            self._render_rows(self.kv1, rows)

    def on_fetch_ids(self):
        if self.ids_fetch_thread and self.ids_fetch_thread.is_alive():
            messagebox.showinfo('Thông báo', 'Đang lấy số định danh, vui lòng chờ...')
            return
        username = (self.ent_ids_user.get() or '').strip() or (self.ent_madv.get().strip() if hasattr(self, 'ent_madv') else '')
        if not username:
            messagebox.showwarning('Thiếu thông tin', 'Vui lòng nhập mã DN (user).')
            return
        try:
            count = int((self.ent_ids_count.get() or '1').strip())
        except ValueError:
            messagebox.showwarning('Giá trị không hợp lệ', 'Số lượng phải là số nguyên.')
            return
        if count <= 0:
            messagebox.showwarning('Giá trị không hợp lệ', 'Số lượng phải lớn hơn 0.')
            return
        node_path = (self.ent_node.get() or '').strip() or shutil.which('node') or shutil.which('node.exe')
        if not node_path:
            messagebox.showwarning('Thiếu Node.js', 'Vui lòng cấu hình đường dẫn node.exe trong tab CSDL.')
            return
        project_dir = os.path.dirname(os.path.abspath(__file__))
        script_path = os.path.join(project_dir, 'fetchIdentifiers.js')
        if not os.path.exists(script_path):
            messagebox.showerror('Thiếu script', f'Không tìm thấy file: {script_path}')
            return
        cmd = [
            node_path,
            script_path,
            '--username', username,
            '--password', username,
            '--count', str(count),
            '--headless', 'false',
        ]
        self.lbl_ids_status.config(text='Đang chạy script Node...')
        thread = threading.Thread(target=self._run_ids_node_process, args=(cmd,), daemon=True)
        self.ids_fetch_thread = thread
        thread.start()

    def on_copy_ids(self, from_event=False):
        if not self.ids_tree:
            return 'break' if from_event else None
        selected = self.ids_tree.selection()
        target_items = selected or self.ids_tree.get_children()
        if not target_items:
            if not from_event:
                messagebox.showinfo('Thông báo', 'Chưa có dữ liệu để copy.')
            return 'break' if from_event else None
        lines = []
        for item_id in target_items:
            values = self.ids_tree.item(item_id, 'values') or ()
            if values:
                lines.append('\t'.join(str(v) for v in values))
        if not lines:
            if not from_event:
                messagebox.showinfo('Thông báo', 'Không có giá trị để copy.')
            return 'break' if from_event else None
        text = '\n'.join(lines)
        try:
            self.clipboard_clear()
            self.clipboard_append(text)
            self.update()
            if not from_event:
                messagebox.showinfo('Đã copy', 'Đã copy các dòng đã chọn.')
        except Exception as exc:
            if not from_event:
                messagebox.showerror('Lỗi', f'Không copy được: {exc}')
        return 'break' if from_event else None

    def _run_ids_node_process(self, cmd):
        try:
            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                stdin=subprocess.PIPE,
                text=True,
                encoding='utf-8',
                errors='ignore',
            )
        except Exception as exc:
            self.after(0, lambda e=exc: self._handle_ids_error(e))
            return

        stderr_lines = []

        def _drain_stderr():
            try:
                for line in proc.stderr:
                    stderr_lines.append(line.rstrip())
            finally:
                try:
                    proc.stderr.close()
                except Exception:
                    pass

        threading.Thread(target=_drain_stderr, daemon=True).start()

        results = []
        error_message = None
        try:
            while True:
                line = proc.stdout.readline()
                if line == '':
                    break
                line = line.strip()
                if not line:
                    continue
                if line.startswith('CAPTCHA'):
                    image_path = None
                    if len(line) > 7:
                        payload = line[8:].strip()
                        if payload:
                            try:
                                meta = json.loads(payload)
                                if isinstance(meta, dict):
                                    image_path = meta.get('image')
                            except Exception:
                                image_path = None
                    try:
                        captcha = self._solve_captcha_auto(image_path)
                    except Exception as exc:
                        try:
                            proc.stdin.write('\n')
                            proc.stdin.flush()
                        except Exception:
                            pass
                        raise exc
                    proc.stdin.write(captcha + '\n')
                    proc.stdin.flush()
                elif line.startswith('RESULT '):
                    payload = line[7:]
                    try:
                        data = json.loads(payload)
                        results = data.get('items', [])
                    except Exception as exc:
                        error_message = f'Không phân tích được kết quả: {exc}'
                elif line.startswith('ERROR '):
                    payload = line[6:]
                    try:
                        data = json.loads(payload)
                        error_message = data.get('message') or payload
                    except Exception:
                        error_message = payload
            proc.wait()
            if error_message:
                raise RuntimeError(error_message)
            if proc.returncode not in (0, None) and not results:
                raise RuntimeError(f'Script kết thúc với mã {proc.returncode}')
            self.after(0, lambda r=results: self._handle_ids_success(r))
        except Exception as exc:
            try:
                proc.kill()
            except Exception:
                pass
            tail = '\n'.join(stderr_lines[-10:])
            if tail:
                exc = RuntimeError(f'{exc}\n{tail}')
            self.after(0, lambda e=exc: self._handle_ids_error(e))

    def _handle_ids_success(self, results):
        self.ids_fetch_thread = None
        if results:
            self.ids_data.extend(results)
            self._refresh_ids_tree()
            self.lbl_ids_status.config(text=f'Đã lấy {len(results)} số định danh.')
        else:
            self.lbl_ids_status.config(text='Không lấy được số định danh nào.')

    def _handle_ids_error(self, error):
        self.ids_fetch_thread = None
        self.lbl_ids_status.config(text='Lỗi khi lấy số định danh.')
        messagebox.showerror('Lỗi', f'Không lấy được số định danh: {error}')

    def _refresh_ids_tree(self):
        if not self.ids_tree:
            return
        self.ids_tree.delete(*self.ids_tree.get_children())
        for idx, item in enumerate(self.ids_data, start=1):
            self.ids_tree.insert('', 'end', values=(idx, item['code'], item['time']))

    def _prompt_captcha(self, image_path=None):
        event = threading.Event()
        result = {'value': None}
        image_path = (image_path or '').strip()
        auto_guess = ''
        if image_path and os.path.exists(image_path):
            auto_guess = self._auto_ocr_captcha(image_path) or ''

        def ask():
            message = 'Nhập mã xác thực đang hiện trong trình duyệt:'
            if image_path:
                message += f'\nẢnh đã được lưu tại:\n{image_path}'
            value = simpledialog.askstring(
                'Mã xác thực',
                message,
                parent=self,
                initialvalue=auto_guess,
            )
            result['value'] = value
            event.set()

        self.after(0, ask)
        event.wait()
        if not result['value']:
            raise RuntimeError('Chưa nhập mã xác thực.')
        return result['value'].strip()

    def _solve_captcha_auto(self, image_path):
        image_path = (image_path or '').strip()
        if not image_path or not os.path.exists(image_path):
            raise RuntimeError('Không nhận được ảnh captcha từ script Node.')
        print(f'[Captcha OCR] Bắt đầu đọc captcha từ: {image_path}')
        code = self._auto_ocr_captcha(image_path)
        if not code:
            raise RuntimeError('Không đọc được mã captcha tự động.')
        return code

    def _run_python_ocr_script(self, script_path, image_path, extra_args=None):
        if not script_path or not os.path.exists(script_path):
            return ''
        python_exec = sys.executable or 'python'
        cmd = [python_exec, script_path]
        if extra_args:
            cmd.extend(extra_args)
        cmd.append(image_path)
        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                encoding='utf-8',
                errors='ignore',
                timeout=25,
            )
        except Exception:
            return ''
        if result.returncode not in (0, None):
            return ''
        return (result.stdout or '').strip()

    def _auto_ocr_captcha(self, image_path):
        if not image_path or not os.path.exists(image_path):
            return ''
        project_dir = os.path.dirname(os.path.abspath(__file__))
        engines = []

        if _ocr_paddle_module is not None:
            engines.append(('paddle', lambda: _ocr_paddle_module.recognize_image(image_path, 'en')))
        if _ocr_easy_module is not None:
            engines.append(('easy', lambda: _ocr_easy_module.recognize_image(image_path, 'en')))

        paddle_script = os.path.join(project_dir, 'ocr_paddle.py')
        easy_script = os.path.join(project_dir, 'ocr_easy.py')

        if os.path.exists(paddle_script):
            engines.append(('paddle-subprocess', lambda: self._run_python_ocr_script(paddle_script, image_path)))
        if os.path.exists(easy_script):
            engines.append(('easy-subprocess', lambda: self._run_python_ocr_script(easy_script, image_path)))

        if not engines:
            print('[Captcha OCR] Không có engine OCR nào khả dụng.')
            return ''

        print(f'[Captcha OCR] Thử OCR với {len(engines)} engine(s): {[name for name, _ in engines]}')
        for engine_name, runner in engines:
            try:
                text = runner() or ''
            except Exception as exc:
                print(f'[Captcha OCR] Engine {engine_name} lỗi: {exc}')
                continue
            cleaned = ''.join(ch for ch in text.strip() if ch.isalnum())
            print(f'[Captcha OCR] Engine {engine_name} raw="{text}" cleaned="{cleaned}"')
            if cleaned:
                print(f'[Captcha OCR] Engine {engine_name} đọc được: {cleaned}')
                return cleaned
        return ''

    def _prepare_rows_from_excel(self, path):
        if load_workbook is None:
            raise RuntimeError('openpyxl chưa được cài đặt.')
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        prepared = []
        skipped = []
        seen_keys = set()
        try:
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row = row or ()
                cells = list(row[:4])
                while len(cells) < 4:
                    cells.append(None)
                mst, ma_hq, so_tk, ngay = cells
                mst = str(mst).strip() if mst not in (None, '') else ''
                ma_hq = str(ma_hq).strip() if ma_hq not in (None, '') else ''
                so_tk = str(so_tk).strip() if so_tk not in (None, '') else ''
                if not mst or not so_tk:
                    skipped.append(f'Hàng {idx}: thiếu MST/Số tờ khai')
                    continue
                ngay_fmt = self._format_mv_date(ngay)
                key = (mst, so_tk, ma_hq, ngay_fmt)
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                prepared.append((mst, so_tk, ma_hq, ngay_fmt))
        finally:
            wb.close()
        return prepared, skipped

    def _move_selected(self, src, dst):
        moved = 0
        for item in src.get_children():
            if src.set(item, 'sel') == CHECK_MARK:
                vals = list(src.item(item, 'values'))
                vals[0] = ''  # clear Sel ở bảng đích
                row_data = self._values_to_row_data(vals)
                dst.insert('', 'end', values=vals)
                src.delete(item)
                if src is self.kv1:
                    self._remove_row_from_cache(row_data)
                if dst is self.kv1:
                    self._add_row_to_cache(row_data)
                moved += 1
        # Cập nhật số lượng và trạng thái checkbox "Tích toàn bộ" sau khi di chuyển
        self._update_count(src)
        self._update_count(dst)
        self._update_select_all_checkbox(src)
        self._update_select_all_checkbox(dst)
        if src is self.kv1 or dst is self.kv1:
            self._refresh_kv1_view()
        return moved

    def on_data_up(self):
        self._move_selected(self.kv2, self.kv1)

    def on_data_down(self):
        self._move_selected(self.kv1, self.kv2)

    def _on_select_all_toggle(self, tree, select_all_var):
        """Xử lý khi click checkbox 'Tích toàn bộ'"""
        print(f"Debug: _on_select_all_toggle called, select_all_var.get() = {select_all_var.get()}")
        items = tree.get_children()
        print(f"Debug: tree has {len(items)} items: {items}")
        
        if select_all_var.get():
            # Tích toàn bộ
            for item in items:
                tree.set(item, 'sel', CHECK_MARK)
                print(f"Debug: Set item {item} to {CHECK_MARK}")
        else:
            # Bỏ tích toàn bộ
            for item in items:
                tree.set(item, 'sel', '')
                print(f"Debug: Set item {item} to empty")
        
        # Force refresh để hiển thị thay đổi
        tree.update()

    def _update_select_all_checkbox(self, tree):
        """Cập nhật trạng thái checkbox 'Tích toàn bộ' dựa trên trạng thái các checkbox riêng lẻ"""
        if not hasattr(tree, 'select_all_var'):
            return
        
        total_items = len(tree.get_children())
        if total_items == 0:
            tree.select_all_var.set(False)
            return
        
        selected_count = 0
        for item in tree.get_children():
            if tree.set(item, 'sel') == CHECK_MARK:
                selected_count += 1
        
        # Cập nhật trạng thái checkbox "Tích toàn bộ"
        if selected_count == 0:
            tree.select_all_var.set(False)
        elif selected_count == total_items:
            tree.select_all_var.set(True)
        else:
            # Trạng thái indeterminate (một số được chọn, một số không)
            tree.select_all_var.set(False)

    def on_choose_folder(self):
        d = filedialog.askdirectory(initialdir=self.output_dir)
        if d:
            self.output_dir = d
            self.lbl_out.config(text=f'Lưu tại: {self.output_dir}')
            self.lbl_out.configure(cursor='hand2', foreground='#0066cc')

    def _open_output_dir(self):
        path = self.output_dir
        if not path:
            messagebox.showwarning('Thiếu thư mục', 'Chưa thiết lập thư mục lưu output.')
            return
        try:
            os.makedirs(path, exist_ok=True)
            if os.name == 'nt':
                os.startfile(path)
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', path])
            else:
                subprocess.Popen(['xdg-open', path])
        except Exception as exc:
            messagebox.showerror('Lỗi', f'Không mở được thư mục output: {exc}')

    def on_choose_node(self):
        path = filedialog.askopenfilename(title='Chọn node.exe', filetypes=[('node.exe', 'node.exe'), ('Tất cả', '*.*')])
        if path:
            self.ent_node.delete(0, tk.END)
            self.ent_node.insert(0, path)

    def on_choose_dshh(self):
        initial = (self.ent_dshh.get() or '').strip() or (self.default_dshh_path or '')
        initialdir = os.path.dirname(initial) if initial else None
        path = filedialog.askopenfilename(
            title='Chọn DSHH.xlsx',
            filetypes=[('Excel', '*.xlsx'), ('Tất cả', '*.*')],
            initialdir=initialdir or None,
        )
        if path:
            self.ent_dshh.delete(0, tk.END)
            self.ent_dshh.insert(0, path)
            self.use_dshh_var.set(True)

    def on_open_dshh_template(self):
        template_path = self.default_dshh_path
        if not template_path or not os.path.isfile(template_path):
            messagebox.showwarning('Thiếu file', f'Không tìm thấy template DSHH tại {template_path or "đường dẫn rỗng"}.')
            return
        try:
            if os.name == 'nt':
                os.startfile(template_path)
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', template_path])
            else:
                subprocess.Popen(['xdg-open', template_path])
        except Exception as exc:
            messagebox.showerror('Lỗi', f'Không mở được file template: {exc}')

    def on_export(self):
        def task():
            try:
                Sqlhost = self._get_sqlhost()
                rows = self.kv2.get_children()
                if not rows:
                    messagebox.showinfo('Thông báo', 'KV2 không có dòng nào để xuất.')
                    return
                outdir = self.output_dir
                prepared_rows = []
                id_list = []
                for item in rows:
                    vals = self.kv2.item(item, 'values')
                    if not vals:
                        continue
                    raw_id = vals[1] if len(vals) > 1 else ''
                    text_id = normalize_dtokhaimdid(raw_id)
                    if not text_id:
                        continue
                    prepared_rows.append((text_id, vals))
                    if text_id not in id_list:
                        id_list.append(text_id)
                if not prepared_rows:
                    messagebox.showinfo('Thong bao', 'KV2 khong co dong hop le de xuat.')
                    return
                bulk_data = get_cd_details_bulk(Sqlhost, id_list)
                if not bulk_data:
                    messagebox.showerror('Loi', 'Khong lay duoc du lieu tu database.')
                    return
                missing = []
                idx = 1
                for text_id, vals in prepared_rows:
                    data = bulk_data.get(text_id)
                    if data is None:
                        missing.append(text_id)
                        continue
                    imex = (vals[2] or '').strip().upper() if len(vals) > 2 else ''
                    if imex == 'X':
                        PrintCD_TKX(text_id, idx, data, outdir)
                    else:
                        PrintCD_TKN(text_id, idx, data, outdir)
                    idx += 1
                if missing:
                    messagebox.showwarning('Thong bao', f"Khong lay duoc du lieu cho ID: {', '.join(missing)}")
                messagebox.showinfo('Hoan thanh', 'Xuat to khai Excel hoan tat!')
            except Exception as e:
                messagebox.showerror('Lỗi', f'Xuất tờ khai lỗi: {e}')
        threading.Thread(target=task, daemon=True).start()

    def on_export_mv(self):
        excel_path = (self.ent_dshh.get() or '').strip() if self.ent_dshh else ''
        use_excel = bool(self.use_dshh_var.get())
        prepared_rows = []
        skipped = []
        selected_items = []

        if use_excel:
            if not excel_path:
                messagebox.showwarning('Thiếu file', 'Chưa chọn đường dẫn file DSHH.xlsx.')
                return
            if not os.path.isfile(excel_path):
                messagebox.showwarning('Thiếu file', f'Không tìm thấy DSHH.xlsx tại: {excel_path}')
                return
            if load_workbook is None:
                messagebox.showerror('Thiếu thư viện', 'Không đọc được file Excel vì chưa cài openpyxl (pip install openpyxl).')
                return

        if not use_excel:
            rows = self.kv2.get_children()
            if not rows:
                messagebox.showinfo('Thông báo', 'KV2 không có dòng nào để tải mã vạch.')
                return
        else:
            rows = []

        outdir = self.output_dir
        os.makedirs(outdir, exist_ok=True)

        node_path = (self.ent_node.get() or '').strip() or shutil.which('node') or shutil.which('node.exe')
        if not node_path:
            messagebox.showwarning('Thiếu Node.js', 'Không tìm thấy Node.js trong PATH. Vui lòng cài Node.js để tải mã vạch.')
            return

        madv = (self.ent_madv.get() or '').strip()
        if not use_excel and not madv:
            messagebox.showwarning('Thiếu MA_DV', 'Vui lòng nhập MA_DV trước khi xuất mã vạch.')
            return

        if use_excel:
            try:
                prepared_rows, skipped = self._prepare_rows_from_excel(excel_path)
            except Exception as exc:
                messagebox.showerror('Đọc DSHH.xlsx', f'Không đọc được file DSHH: {exc}')
                return
        else:
            selected_items = [item for item in rows if self.kv2.set(item, 'sel') == CHECK_MARK]
            items_to_use = selected_items or list(rows)
            seen_keys = set()
            for item in items_to_use:
                vals = self.kv2.item(item, 'values')
                if not vals:
                    skipped.append('dòng trống')
                    continue

                so_tk = (vals[3] or '').strip() if len(vals) > 3 else ''
                ma_hq = (vals[5] or '').strip() if len(vals) > 5 else ''
                raw_ngay = (vals[6] or '').strip() if len(vals) > 6 else ''
                batch_id = (vals[1] or '').strip() if len(vals) > 1 else so_tk or '?'

                if not so_tk:
                    skipped.append(f'{batch_id}: thiếu số tờ khai')
                    continue

                key = (so_tk, ma_hq, raw_ngay)
                if key in seen_keys:
                    continue
                seen_keys.add(key)

                prepared_rows.append((madv, so_tk, ma_hq, self._format_mv_date(raw_ngay)))

        if not prepared_rows:
            messagebox.showwarning('Thông báo', 'Không có dòng hợp lệ để tạo batch mã vạch.')
            return

        csv_path = os.path.join(outdir, 'mv_batch.csv')
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['MaDoanhNghiep', 'SoToKhai', 'MaHQ', 'NgayToKhai'])
            writer.writerows(prepared_rows)

        project_dir = os.path.dirname(os.path.abspath(__file__))
        script_path = os.path.join(project_dir, 'scrapeBarcodeContainer.aspx.js')
        if not os.path.exists(script_path):
            messagebox.showerror('Thiếu script', f'Không tìm thấy file: {script_path}')
            return

        out_pattern = os.path.join(outdir, 'MV_{SoToKhai}.pdf')
        state_file = os.path.join(outdir, 'mv_state.json')
        log_path = os.path.join(outdir, 'mv_log.txt')
        easy_script = os.path.join(project_dir, 'ocr_easy.py')

        args = [
            node_path,
            script_path,
            '--batch', csv_path,
            '--out-pattern', out_pattern,
            '--save-state', state_file,
            '--log', log_path,
            '--auto-only', 'true',
            '--ocr-tries', '4',
        ]
        easy_ok = getattr(self, '_easy_available', None)
        if easy_ok is None:
            easy_ok = _detect_easy_available()
            self._easy_available = easy_ok
        if easy_ok and not os.path.exists(easy_script):
            easy_ok = False
            self._easy_available = False

        chosen_engine = None
        chosen_script = None
        if easy_ok:
            chosen_engine = 'easy'
            chosen_script = easy_script

        if chosen_engine and chosen_script:
            python_exec = sys.executable or 'python'
            args.extend([
                '--ocr-engine', chosen_engine,
                '--ocr-python', python_exec,
                '--ocr-script', chosen_script,
            ])
            if chosen_engine in ('paddle', 'easy'):
                args.extend(['--no-fallback', 'true'])
        else:
            if not getattr(self, '_ocr_warned', False):
                messagebox.showinfo(
                    'Thông báo',
                    'Không tìm thấy EasyOCR trong môi trường Python hiện tại. Sẽ sử dụng Tesseract OCR.'
                )
                self._ocr_warned = True
        if os.path.exists(state_file):
            args.extend(['--load-state', state_file])

        exported_count = len(prepared_rows)
        skipped_preview = ', '.join(skipped[:5]) if skipped else ''
        selected_info = len(selected_items)

        def run_mv_task():
            try:
                self.after(0, lambda: self._update_status(self.kv2, 'loading'))
                result = subprocess.run(
                    args,
                    cwd=project_dir,
                    capture_output=True,
                    text=True,
                    encoding='utf-8',
                    errors='ignore',
                    timeout=600,
                )
                if result.returncode == 0:
                    self.after(0, lambda: self._update_status(self.kv2, 'completed'))
                    summary_lines = [
                        f'Đã gửi {exported_count} dòng tới Node.js.',
                        f'File batch: {os.path.basename(csv_path)}',
                    ]
                    if selected_info:
                        summary_lines.append(f'Sử dụng {selected_info} dòng đã tích.')
                    if skipped:
                        summary_lines.append(
                            f'Bỏ qua {len(skipped)} dòng: {skipped_preview}{", ..." if len(skipped) > 5 else ""}'
                        )
                    self.after(0, lambda: messagebox.showinfo('Hoàn thành', '\n'.join(summary_lines)))
                    self.after(3000, lambda: self._update_status(self.kv2, 'clear'))
                else:
                    stderr = (result.stderr or '').strip()
                    snippet = '\n'.join(stderr.splitlines()[:5]) if stderr else 'Không có log lỗi.'
                    self.after(0, lambda: self._update_status(self.kv2, 'error'))
                    self.after(0, lambda: messagebox.showerror('Lỗi Node.js', f'Node trả về mã {result.returncode}.\n{snippet}'))
                    self.after(5000, lambda: self._update_status(self.kv2, 'clear'))
            except subprocess.TimeoutExpired:
                self.after(0, lambda: self._update_status(self.kv2, 'error'))
                self.after(0, lambda: messagebox.showerror('Timeout', 'Node.js không phản hồi trong 10 phút.'))
                self.after(5000, lambda: self._update_status(self.kv2, 'clear'))
            except Exception as exc:
                self.after(0, lambda: self._update_status(self.kv2, 'error'))
                self.after(0, lambda: messagebox.showerror('Lỗi', f'Xuất Mã Vạch lỗi: {exc}'))
                self.after(5000, lambda: self._update_status(self.kv2, 'clear'))

        threading.Thread(target=run_mv_task, daemon=True).start()


if __name__ == '__main__':
    app = PrintApp()
    app.mainloop()
